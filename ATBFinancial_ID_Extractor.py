"""
Netomi ATB chatbot — full UI flow including OTP + conversationId capture,
followed automatically by an API-driven multi-turn conversation runner.

UI phase (per row of queries.xlsx):
    1. Open https://demo.netomi.com/web/ATB/ATBFinancial.html?ref=ncw
    2. If #netomiChatWindow iframe isn't already attached, click "Let's talk!"
    3. Click "Agree" disclaimer inside the iframe if it appears
    4. Wait for bot greeting "How can I assist you today?" to be visible
    5. Type the query (human-paced), press Enter
    6. Wait for bot's response: <div class="title">Log In</div> card
    7. Click it — a new window/popup opens (the ATB login page)
    8. Wait for #username, type LoginID from column B
    9. Click #main-button (Continue)
   10. Wait for #password, type Password from column C
   11. Click #signOnButton (Log in)
   12. Wait for #passcode to be visible (server has emailed the OTP)
   13. Open mailinator.com/v4/public/inboxes.jsp?to=<LoginID>, grab a fresh OTP
   14. Close the mailinator tab, type the OTP into #passcode
   15. Click #sign-on (Verify)
   16. Write conversationId + userId (captured from the widget's webhook POSTs)
       into columns D and E.

API phase (runs automatically after the UI phase):
    - Imports atb_api_runner and invokes run_sync(...) with the active test sheet
      (ATBFinancial_Staging.xlsx for STAGING, ATBFinancial_Dev.xlsx for DEV).
    - Uses API_WORKERS for parallel TestIDs (independent of UI_WORKERS).
    - See atb_api_runner.py for the per-turn pipeline and schema details.

queries.xlsx schema:
    A = query
    B = LoginID
    C = Password
    D = conversationId   (written on success; "ERROR: ..." on failure)
    E = userId           (written on success alongside conversationId)

Usage:
    pip install playwright openpyxl httpx
    python -m playwright install chromium firefox webkit
    python netomi_conversation_extractor.py
"""
from __future__ import annotations
import asyncio
import json
import os
import random
import re
import sys
import tempfile
import time
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook, load_workbook
from playwright.async_api import (
    async_playwright,
    Browser,
    BrowserContext,
    Frame,
    Page,
    Playwright,
    Request,
    TimeoutError as PWTimeout,
)

# =============================================================================
# ALL TUNABLE CONFIG LIVES IN THIS BLOCK. Edit any value below to change
# behavior; nothing below this block should ever need to be touched.
# =============================================================================

# ---------- FILE PATHS ----------
# Input/output workbook for the UI phase (user pool: query, LoginID, Password,
# then D=conversationId, E=userId after UI capture). Single file shared across
# environments — D/E get cleared between STAGING and DEV runs so each captures
# its own fresh IDs.
INPUT_XLSX: str = "queries.xlsx"
# Per-environment test sheets. The STAGING sheet is processed when
# BOT_ENVIRONMENT="STAGING" and the DEV sheet when BOT_ENVIRONMENT="DEV".
# When RUN_BOTH_ENVIRONMENTS=True, STAGING runs first against its sheet, then
# DEV against its sheet — no archiving or clearing of test sheets is needed
# because the two are physically separate files.
# Both sheets follow the same schema:
#   A=Query, B=TestID, C=Payload, D=conversationId,
#   E=userId, F=LoginID, G=Responses
TEST_SHEET_STAGING_XLSX: str = "ATBFinancial_Staging.xlsx"
TEST_SHEET_DEV_XLSX: str = "ATBFinancial_Dev.xlsx"

# ---------- PHASE CONTROL ----------
# Controls what happens before the API phase runs. Choose one:
#   "REFRESH_ALL"  → wipe every row's D/E in queries.xlsx, then re-capture fresh IDs
#                    for all rows via the UI flow. Use this when you want a clean
#                    set of logins each run.
#   "FILL_MISSING" → only run the UI flow for rows whose D/E are blank or marked
#                    ERROR. Rows with valid conversationId/userId are left
#                    untouched. Use this to resume after a partial failure.
#   "SKIP"         → don't open a browser at all. Go straight to the API runner
#                    using whatever IDs already exist in queries.xlsx. Useful
#                    when you already have a populated user pool and just want
#                    to (re)run the API flow.
UI_PHASE_MODE: str = "REFRESH_ALL"

# ---------- BOT ENVIRONMENT ----------
# Selects which ATB chatbot deployment to test against. Each preset
# configures the demo URL and the login-click style for that environment.
#   "STAGING" = official production bot at ATBFinancial.html?ref=ncw.
#               Login is a card with a "Log In" title; clicking it opens a
#               popup window with the login form.
#   "DEV"     = development bot at ATBFinancialDEV.html. Login is an inline
#               <a> link inside a .bot-agent-response paragraph (text usually
#               "ATB login"); clicking it opens login-uat.atb.com in a new tab.
# This affects DEMO_URL and which login-element waiter is used. When
# LETS_TALK_MODE is "DEFAULT" (recommended), this also picks the right
# Let's-talk click behavior (STAGING → NEVER, DEV → ALWAYS).
BOT_ENVIRONMENT: str = "STAGING"

# When True, the script runs STAGING first against TEST_SHEET_STAGING_XLSX,
# then clears the queries.xlsx D/E columns (so DEV captures fresh IDs),
# switches BOT_ENVIRONMENT to DEV, and runs the entire pipeline again
# against TEST_SHEET_DEV_XLSX. The two test sheets are independent files
# you provide before running — no archiving or test-sheet clearing
# happens because they're physically separate. The user-set BOT_ENVIRONMENT
# above is overridden when this flag is on; both environments run regardless.
RUN_BOTH_ENVIRONMENTS: bool = False

# ---------- BROWSER / SESSION (UI phase only) ----------
# Which browser to drive:
#   "chrome"   → Google Chrome (channel). Best stealth; requires Chrome installed.
#   "chromium" → Playwright's bundled Chromium.
#   "webkit"   → WebKit engine (behind Safari). NOT the actual Safari.app.
#   "firefox"  → Playwright's bundled Firefox.
#   "opera"    → Real Opera via executable_path.
BROWSER: str = "chrome"
OPERA_EXEC_PATH: str = "/Applications/Opera.app/Contents/MacOS/Opera"
HEADLESS: bool = True
HUMAN_LIKE: bool = True
FULL_SCREEN_VIEWPORT: dict = {"width": 1920, "height": 1080}

# Controls how the script handles the "Let's talk!" trigger that opens the
# chat widget. The recommended setting is "DEFAULT" — it picks the right
# behavior automatically based on BOT_ENVIRONMENT. Override only if your
# specific deployment behaves differently from the norm.
#   "DEFAULT" = pick automatically based on BOT_ENVIRONMENT:
#                 STAGING → NEVER (chat auto-opens; clicking would close it)
#                 DEV     → ALWAYS (chat needs an explicit click to open)
#   "AUTO"    = wait briefly for the iframe to auto-attach. If it doesn't
#               appear within a few seconds, click "Let's talk!" and try
#               again. Use when behavior is inconsistent across runs.
#   "ALWAYS"  = always click "Let's talk!" first, then wait for the iframe.
#               Use when your environment requires the click every time.
#   "NEVER"   = never click; just wait for the iframe to attach on its own.
#               Use when the chat opens automatically and clicking would
#               close an already-open widget.
LETS_TALK_MODE: str = "DEFAULT"

# ---------- PARALLELISM ----------
# UI_WORKERS:   parallel browser contexts in the UI phase. Keep this
#               conservative — each worker is a full real browser.
# API_WORKERS:  parallel TestID workers in the API phase. Can comfortably go
#               higher than UI_WORKERS since each one is just HTTP requests.
#               Within one TestID, turns are still strictly sequential
#               (multi-turn conversation order is preserved); parallelism is
#               at the TestID level.
# MAX_WORKERS:  hard safety cap clamped onto both worker counts.
UI_WORKERS: int = 1
API_WORKERS: int = 1
MAX_WORKERS: int = 50

# ---------- UI ↔ API PIPELINING ----------
# False (default) → run all UI rows first, then run all API TestIDs at the end.
#                   API uses API_WORKERS; UI uses UI_WORKERS.
# True            → after every UI batch of UI_WORKERS rows, immediately run
#                   the API flow for any TestIDs that just became ready
#                   (skipping ones already processed in earlier batches). API
#                   uses UI_WORKERS in this mode (matches the batch size).
#                   This pipelines the two phases so you don't have to wait
#                   for all UI to finish before any API begins. With
#                   UI_WORKERS=1, "batch" means a single row.
PARALLEL_PROCESSING: bool = True

# ---------- UI RETRY / TIMEOUT TUNING ----------
# Attempts per query within one pass over the rows.
INNER_RETRIES: int = 2
# Max number of times we'll re-walk the rows that still don't have IDs.
MAX_PASSES: int = 10
# Per-attempt total timeout (from navigation to password fill).
ATTEMPT_TIMEOUT_S: int = 120
# Safety wait after the Verify click for the webhook POST that carries
# conversationId. In practice it's usually resolved from the greeting POST already.
CONV_ID_WAIT_S: int = 30
# Breather between passes before re-walking rows that still need work.
INTER_PASS_DELAY_S: float = 2.0
# Backoff base for inner attempt retries. Sleep between attempts = base * attempt_number.
RETRY_BACKOFF_BASE_S: float = 1.5

# After typing the user's query and pressing Enter, the bot is supposed to
# reply with a login card (STAGING) or a login link (DEV). If that response
# doesn't appear within LOGIN_LINK_WAIT_TIMEOUT_S, we re-send the same
# query up to LOGIN_LINK_QUERY_RETRIES times before failing the row.
# Total time budget for this step ≈ (RETRIES + 1) * WAIT_TIMEOUT_S.
#
# LOGIN_LINK_POLL_INTERVAL_S is how often the script re-checks the chat
# frame for a login element while waiting. The wait is a continuous polling
# loop — every poll interval, all login-target candidate selectors are
# checked. Lower value = catches the element faster once it appears (so we
# click and proceed sooner), at the cost of slightly more CPU. 2s is a
# good balance for a Vue widget.
#
# Re-sends happen IMMEDIATELY when the per-attempt wait window elapses, with
# no extra sleep — polling continues right up until the moment of re-send,
# so a slow bot reply never gets missed in the gap.
LOGIN_LINK_WAIT_TIMEOUT_S: int = 25
LOGIN_LINK_QUERY_RETRIES: int = 3
LOGIN_LINK_POLL_INTERVAL_S: float = 2.0

# When True, save a full-page PNG screenshot of the browser context whenever
# an attempt of process_query fails. Lets you see exactly what state the
# widget was in at the moment of breakage. Files land in UI_SCREENSHOT_DIR
# (created automatically if missing). Filenames include row, login_id,
# attempt number, and a timestamp for uniqueness.
UI_SCREENSHOT_ON_FAILURE: bool = True
UI_SCREENSHOT_DIR: str = "ui_error_screenshots"

# ---------- OTP / MAILINATOR ----------
# Two ways to gate "is this OTP fresh enough":
#   OTP_REQUIRE_JUST_NOW=True  → only accept OTPs whose mailinator "received"
#                                column literally contains "just now". Strictest.
#                                Rejects "30 seconds ago", "1 minute ago", etc.
#   OTP_REQUIRE_JUST_NOW=False → fall back to a minutes-based comparison
#                                using OTP_FRESHNESS_MINUTES (older behavior).
OTP_REQUIRE_JUST_NOW: bool = True
# Only consulted when OTP_REQUIRE_JUST_NOW=False.
OTP_FRESHNESS_MINUTES: int = 1
# Hard cap on the number of mailinator page loads (initial load + reloads
# combined) before giving up. The fetch returns early once this is hit,
# even if OTP_FETCH_TIMEOUT_S hasn't elapsed. With OTP_POLL_INTERVAL_S=10s
# and OTP_MAX_ATTEMPTS=5 the OTP loop wraps in ~50s instead of dragging
# out to the full 180s timeout. Raise if your bot sometimes sends OTP
# emails after a long delay.
OTP_MAX_ATTEMPTS: int = 5
# Total time budget (seconds) — outer cap, applied alongside OTP_MAX_ATTEMPTS.
# Whichever limit hits first wins.
OTP_FETCH_TIMEOUT_S: int = 180
# Reload interval while waiting for a fresh OTP to arrive.
OTP_POLL_INTERVAL_S: int = 10

# ---------- API PHASE TUNING ----------
# False → TestIDs are paired sequentially with queries.xlsx rows (TestID k ← user k).
# True  → read D/E/F from the test sheet (pre-filled by you) and validate them.
MANUAL_ASSIGNMENT: bool = False
# Turn-level retries for the full POST+GET cycle. Total attempts per turn = this + 1.
API_TURN_RETRIES: int = 2
# Per-HTTP-call timeout (POST and each GET attempt).
API_TURN_TIMEOUT_S: float = 60.0
# Delay between sequential turns of one TestID. Set 0 to disable.
API_INTER_TURN_DELAY_S: float = 0.5
# GET polling inside one turn: max attempts (suffix flips each attempt) and delay.
API_GET_MAX_ATTEMPTS: int = 20
API_GET_RETRY_DELAY_S: float = 5.0

# When True, after the main pipeline finishes (UI + API, whether sequential or
# pipelined), the script does ONE more pass that rescans the test sheet for
# TestIDs whose Response column has ERROR: or SKIPPED: markers and re-runs
# them using the original user (conv_id, user_id, LoginID) stored in their
# D/E/F columns. The failure markers are cleared first so the retry writes
# fresh values. TestIDs without an original user in D/E are skipped.
RETRY_FAILED_API_TURNS: bool = False

# Per-environment Netomi bot reference IDs (sent as the x-bot-ref-id header
# on every POST). Each environment is registered as a separate bot on
# Netomi's side, so the wrong ID will silently produce empty/nothing
# responses. Both runs (single-env and RUN_BOTH_ENVIRONMENTS) pick the right
# value automatically based on BOT_ENVIRONMENT.
STAGING_BOT_REF_ID: str = "fab252f5-dd1b-4be7-9a39-898585ddb769"
DEV_BOT_REF_ID: str = "fe3b08f5-650b-490d-988c-34d0e6ecdf65"

# =============================================================================
# END OF TUNABLES.
# =============================================================================
STAGING_DEMO_URL = "https://demo.netomi.com/web/ATB/ATBFinancial.html?ref=ncw"
DEV_DEMO_URL = "https://demo.netomi.com/web/ATB/ATBFinancialDEV.html"
DEMO_URL = STAGING_DEMO_URL if BOT_ENVIRONMENT == "STAGING" else DEV_DEMO_URL
# Active test sheet path — picked from the per-environment files in the
# config block based on BOT_ENVIRONMENT. _set_bot_environment() updates this
# in place when switching between STAGING and DEV during a both-envs run.
TEST_SHEET_XLSX = (
    TEST_SHEET_STAGING_XLSX if BOT_ENVIRONMENT == "STAGING" else TEST_SHEET_DEV_XLSX
)
MAILINATOR_URL_TEMPLATE = "https://mailinator.com/v4/public/inboxes.jsp?to={login_id}"
# Substring match on the outgoing POST we're scraping for conversationId.
WEBHOOK_URL_SUBSTR = "chatapps-us.netomi.com/api/v1/webhook-message"
CHAT_IFRAME_ID = "netomiChatWindow"
CHROMIUM_FAMILY = {"chrome", "chromium", "opera"}

# queries.xlsx column indexes (0-based within the in-memory rows list).
COL_QUERY = 0
COL_LOGIN_ID = 1
COL_PASSWORD = 2
COL_CONV_ID = 3
COL_USER_ID = 4
NUM_COLUMNS = 5
DEFAULT_HEADER = ["query", "LoginID", "Password", "conversationId", "userId"]

# Selectors
LETS_TALK_SELECTORS = [
    # OUTER <button> FIRST. The widget is a real <button id=
    # "support-icon-container"> — the proper click target with the bound
    # event handler. Clicking the button directly is more reliable than
    # clicking the inner text/div and relying on event bubbling.
    ("css", "button#support-icon-container"),
    ("css", "#support-icon-container"),
    ("css", "button.support-icon-container"),
    ("css", "button[aria-label='chat']"),
    # Text-based fallbacks for deployments that don't match the button
    # selectors above (Playwright's get_by_text returns the most specific
    # containing element, which still bubbles its click up to the button).
    ("text", "Let's talk!"),
    ("regex", r"(?i)let.?s\s*talk"),
    # Inner-element class fallbacks for the oldest deployments.
    ("css", ".support-text"),
    ("css", "div.support-icon.text-icon-right"),
    ("css", ".support-icon"),
]

# ---------- stealth ----------
STEALTH_COMMON = r"""
(() => {
  try { Object.defineProperty(navigator, 'webdriver', { get: () => undefined }); } catch (e) {}
})();
"""

STEALTH_CHROMIUM = r"""
(() => {
  try {
    Object.defineProperty(navigator, 'plugins', {
      get: () => {
        const p = [
          { name: 'PDF Viewer', filename: 'internal-pdf-viewer', description: 'Portable Document Format' },
          { name: 'Chrome PDF Viewer', filename: 'internal-pdf-viewer', description: 'Portable Document Format' },
          { name: 'Chromium PDF Viewer', filename: 'internal-pdf-viewer', description: 'Portable Document Format' },
          { name: 'Microsoft Edge PDF Viewer', filename: 'internal-pdf-viewer', description: 'Portable Document Format' },
          { name: 'WebKit built-in PDF', filename: 'internal-pdf-viewer', description: 'Portable Document Format' },
        ];
        p.item = i => p[i];
        p.namedItem = n => p.find(x => x.name === n);
        return p;
      }
    });
  } catch (e) {}
  try { Object.defineProperty(navigator, 'languages', { get: () => ['en-US', 'en'] }); } catch (e) {}
  try { Object.defineProperty(navigator, 'platform', { get: () => 'MacIntel' }); } catch (e) {}
  try { Object.defineProperty(navigator, 'hardwareConcurrency', { get: () => 8 }); } catch (e) {}
  try { Object.defineProperty(navigator, 'deviceMemory', { get: () => 8 }); } catch (e) {}
  if (!window.chrome) { window.chrome = {}; }
  window.chrome.runtime = window.chrome.runtime || {};
  window.chrome.loadTimes = window.chrome.loadTimes || function () {};
  window.chrome.csi = window.chrome.csi || function () {};
  try {
    const origQuery = window.navigator.permissions && window.navigator.permissions.query;
    if (origQuery) {
      window.navigator.permissions.query = (parameters) => (
        parameters && parameters.name === 'notifications'
          ? Promise.resolve({ state: Notification.permission })
          : origQuery(parameters)
      );
    }
  } catch (e) {}
  try { delete window.cdc_adoQpoasnfa76pfcZLmcfl_Array; } catch (e) {}
  try { delete window.cdc_adoQpoasnfa76pfcZLmcfl_Promise; } catch (e) {}
  try { delete window.cdc_adoQpoasnfa76pfcZLmcfl_Symbol; } catch (e) {}
})();
"""

# ---------- human-like helpers ----------
async def _rand_sleep(lo: float, hi: float) -> None:
    if HUMAN_LIKE:
        await asyncio.sleep(random.uniform(lo, hi))

async def human_click(locator) -> None:
    if HUMAN_LIKE:
        try:
            await locator.hover()
        except Exception:
            pass
        await _rand_sleep(0.15, 0.4)
    await locator.click()

async def human_type(locator, text: str) -> None:
    await locator.click()
    await _rand_sleep(0.15, 0.35)
    if HUMAN_LIKE:
        await locator.type(text, delay=random.randint(50, 110))
    else:
        await locator.fill(text)
    await _rand_sleep(0.2, 0.5)

# ---------- xlsx helpers ----------
def _stringify_cell(v) -> str:
    """Normalize a cell value to a string. Integers stay integers (no ".0")."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        return str(v)
    return str(v)

def load_xlsx(path: Path) -> tuple[list[list[str]], list[str]]:
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    # Some xlsx files (especially ones generated outside Excel) carry a malformed
    # "active sheet" reference. openpyxl strips it and wb.active returns None.
    # Fall back to the first worksheet in that case.
    if ws is None:
        sheets = wb.worksheets
        if not sheets:
            wb.close()
            raise RuntimeError(f"{path} has no worksheets")
        ws = sheets[0]
    rows: list[list[str]] = []
    header: list[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        row_list = [_stringify_cell(c) for c in row]
        if i == 0:
            header = row_list
        else:
            rows.append(row_list)
    wb.close()
    if not header and not rows:
        raise RuntimeError(f"{path} is empty")
    # Pad each data row to the expected width.
    for r in rows:
        while len(r) < NUM_COLUMNS:
            r.append("")
    return rows, header

def _save_xlsx_once(path: Path, header: list[str], rows: list[list[str]]) -> None:
    """Single atomic save attempt."""
    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for r in rows:
        ws.append(r)

    fd, tmp_path = tempfile.mkstemp(
        prefix=path.stem + "_", suffix=".tmp", dir=str(path.parent)
    )
    os.close(fd)  # openpyxl opens by path; close the mkstemp handle first.
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, path)
    except Exception:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise
    finally:
        wb.close()


def save_xlsx(path: Path, header: list[str], rows: list[list[str]]) -> None:
    """
    Atomic xlsx save with retry on transient IO errors.

    If you have the xlsx open in Excel (or another app that briefly takes
    a write lock), the os.replace can momentarily hit a PermissionError /
    OSError. We retry up to 5 times with short backoff so a sub-second
    contention window doesn't surface as a hard write failure. If the file
    stays locked beyond the retry budget, the original error is raised so
    the caller's normal error handling kicks in.

    Behavior with the file open in Excel on macOS: usually no contention
    at all (APFS lets us rename over the file even while it's open). The
    retries are insurance for the edge cases. On Windows, where mandatory
    locking is the norm, the retries cover the typical "file in use by
    another process" scenario as long as the user releases it within ~5s.
    """
    max_retries = 5
    backoff_s = 0.5
    last_err: Optional[Exception] = None
    for attempt in range(max_retries + 1):
        try:
            _save_xlsx_once(path, header, rows)
            return
        except (PermissionError, OSError) as e:
            # FileNotFoundError isn't a lock contention issue — propagate it.
            if isinstance(e, FileNotFoundError):
                raise
            last_err = e
            if attempt < max_retries:
                wait = backoff_s * (attempt + 1)
                print(
                    f"[WARN] save_xlsx({path.name}) hit IO error on attempt "
                    f"{attempt + 1}/{max_retries + 1}: {e}; retrying in "
                    f"{wait:.1f}s. (If you have the file open in Excel and "
                    f"this keeps happening, close it.)",
                    file=sys.stderr,
                )
                time.sleep(wait)
    # Exhausted retries.
    if last_err is not None:
        raise last_err

def needs_work(row: list[str]) -> bool:
    """
    A row still needs processing if EITHER column D (conversationId) or column E
    (userId) is missing / empty / previously errored. Legacy "OK" markers count as
    needs-work too so older files are reprocessed cleanly. The "ERROR:" check
    requires the trailing colon so an ID that incidentally starts with the
    letters E-R-R-O-R isn't misclassified as a failed row.
    """
    def _bad(val: str) -> bool:
        v = (val or "").strip()
        return not v or v.startswith("ERROR:") or v == "OK"
    if len(row) <= COL_USER_ID:
        return True
    return _bad(row[COL_CONV_ID]) or _bad(row[COL_USER_ID])

def normalize_header(header: list[str]) -> list[str]:
    """Ensure the header has exactly NUM_COLUMNS expected labels."""
    out = list(header) if header else []
    while len(out) < NUM_COLUMNS:
        out.append(DEFAULT_HEADER[len(out)])
    return out[:NUM_COLUMNS]

# ---------- selector helpers (main page) ----------
async def find_lets_talk(page: Page, timeout_ms: int = 30000):
    """
    The Netomi widget loads its 'Let's talk!' trigger inside its own iframe
    (separate from the chat conversation iframe that opens later). So we walk
    every frame on the page — main + every nested iframe — and return whichever
    frame has the trigger visible. page.frames refreshes on each outer pass so
    iframes that appear partway through the wait window get picked up too.
    """
    deadline = time.monotonic() + timeout_ms / 1000
    last_err: Optional[Exception] = None
    probe_timeout_ms = 800  # per selector per frame
    while time.monotonic() < deadline:
        for frame in page.frames:
            for kind, val in LETS_TALK_SELECTORS:
                try:
                    if kind == "text":
                        loc = frame.get_by_text(val, exact=True).first
                    elif kind == "regex":
                        loc = frame.get_by_text(re.compile(val)).first
                    else:
                        loc = frame.locator(val).first
                    await loc.wait_for(state="visible", timeout=probe_timeout_ms)
                    return loc
                except Exception as e:
                    last_err = e
                    continue
        await asyncio.sleep(0.3)
    raise RuntimeError(
        f"Could not find 'Let's talk!' trigger in any frame: {last_err}"
    )

# ---------- browser + context construction ----------
async def launch_browser(pw: Playwright) -> Browser:
    if BROWSER in CHROMIUM_FAMILY:
        args = [
            "--disable-blink-features=AutomationControlled",
            "--disable-features=IsolateOrigins,site-per-process",
            "--disable-site-isolation-trials",
            "--no-de    fault-browser-check",
            "--no-first-run",
        ]
        if not HEADLESS:
            args.insert(0, "--start-maximized")
        if BROWSER == "chrome":
            return await pw.chromium.launch(
                channel="chrome",
                headless=HEADLESS,
                args=args,
                ignore_default_args=["--enable-automation"],
            )
        elif BROWSER == "chromium":
            return await pw.chromium.launch(
                headless=HEADLESS,
                args=args,
                ignore_default_args=["--enable-automation"],
            )
        elif BROWSER == "opera":
            if not os.path.exists(OPERA_EXEC_PATH):
                raise RuntimeError(
                    f"Opera not found at {OPERA_EXEC_PATH}. "
                    f"Edit OPERA_EXEC_PATH at the top of the script."
                )
            return await pw.chromium.launch(
                executable_path=OPERA_EXEC_PATH,
                headless=HEADLESS,
                args=args,
                ignore_default_args=["--enable-automation"],
            )
    elif BROWSER == "webkit":
        return await pw.webkit.launch(headless=HEADLESS)
    elif BROWSER == "firefox":
        return await pw.firefox.launch(headless=HEADLESS)
    raise ValueError(
        f"Unknown BROWSER value: {BROWSER!r}. "
        f"Use one of: chrome, chromium, webkit, firefox, opera"
    )

async def new_context(browser: Browser) -> BrowserContext:
    use_no_viewport = (BROWSER in CHROMIUM_FAMILY) and (not HEADLESS)
    kwargs = dict(
        locale="en-US",
        extra_http_headers={"Accept-Language": "en-US,en;q=0.9"},
    )
    if use_no_viewport:
        kwargs["no_viewport"] = True
    else:
        kwargs["viewport"] = FULL_SCREEN_VIEWPORT
    if BROWSER == "chromium" and HEADLESS:
        kwargs["user_agent"] = (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/139.0.0.0 Safari/537.36"
        )
    context = await browser.new_context(**kwargs)
    await context.add_init_script(STEALTH_COMMON)
    if BROWSER in CHROMIUM_FAMILY:
        await context.add_init_script(STEALTH_CHROMIUM)
    return context

# ---------- chat-iframe steps ----------
AGREE_SELECTORS_CSS = [
    '.terms-and-conditions-actions .card-button:has-text("Agree")',
    '.terms-and-conditions-container .card-button:has-text("Agree")',
    '[aria-modal="true"] [role="button"]:has-text("Agree")',
    '[role="dialog"] [role="button"]:has-text("Agree")',
    'div.card-button:has-text("Agree")',
    '[role="button"]:has-text("Agree")',
]
MODAL_PROBE_SELECTORS = [
    '.terms-and-conditions-container',
    '[role="dialog"][aria-modal="true"]',
]

async def _try_click_cascade(loc) -> bool:
    """Try six click strategies in order. Returns True on the first that doesn't raise."""
    try:
        await loc.click(timeout=2500)
        return True
    except Exception:
        pass
    try:
        await loc.scroll_into_view_if_needed(timeout=1500)
        await loc.click(force=True, timeout=2000)
        return True
    except Exception:
        pass
    try:
        await loc.evaluate("el => el.click()")
        return True
    except Exception:
        pass
    try:
        await loc.evaluate(
            "el => el.dispatchEvent(new MouseEvent('click', "
            "{view: window, bubbles: true, cancelable: true, button: 0}))"
        )
        return True
    except Exception:
        pass
    try:
        await loc.focus()
        await loc.press("Enter")
        return True
    except Exception:
        pass
    try:
        await loc.focus()
        await loc.press(" ")
        return True
    except Exception:
        pass
    return False

async def _modal_still_present(scopes) -> bool:
    for scope in scopes:
        for sel in MODAL_PROBE_SELECTORS:
            try:
                if await scope.locator(sel).first.is_visible():
                    return True
            except Exception:
                continue
    return False

async def click_agree(page: Page, chat_frame, timeout_ms: int = 20000) -> bool:
    deadline = time.monotonic() + timeout_ms / 1000

    def _scopes() -> list:
        out = [chat_frame]
        for frame in page.frames:
            out.append(frame)
        return out

    while time.monotonic() < deadline:
        for scope in _scopes():
            for sel in AGREE_SELECTORS_CSS:
                try:
                    loc = scope.locator(sel).first
                    await loc.wait_for(state="visible", timeout=700)
                except Exception:
                    continue
                if await _try_click_cascade(loc):
                    await asyncio.sleep(0.6)
                    if not await _modal_still_present(_scopes()):
                        return True
        await asyncio.sleep(0.3)
    return False

# Greeting phrasings the bot might use. Edit/extend if your environment ships
# a greeting that doesn't match any of these. Matching is case-insensitive and
# uses re.search, so partial matches inside a longer message work.
GREETING_PATTERNS = [
    r"(?i)how\s+can\s+I\s+(?:assist|help)\s+you",
    r"(?i)how\s+may\s+I\s+(?:assist|help)\s+you",
    r"(?i)what\s+can\s+I\s+(?:do|help)\s+for\s+you",
    r"(?i)hi[\s,.!]+(?:how|what)",
    r"(?i)hello[\s,.!]+(?:how|what)",
    r"(?i)welcome",
]


async def wait_for_greeting(chat_frame, timeout_ms: int = 30000) -> None:
    """
    Wait until the chat is ready for the user to type. Returns successfully on
    EITHER signal:
      - One of the GREETING_PATTERNS appears as visible text in the chat frame.
      - The chat input textarea becomes visible (the strongest readiness signal,
        and the one that actually matters since we type into it next).
    Different deployments (prod vs DEV vs different bots on the same widget)
    have different greeting text — relying on textarea visibility means we
    don't need to know the exact phrasing.
    """
    deadline = time.monotonic() + timeout_ms / 1000
    last_err: Optional[Exception] = None
    while time.monotonic() < deadline:
        # Try every greeting pattern.
        for pattern in GREETING_PATTERNS:
            try:
                loc = chat_frame.get_by_text(re.compile(pattern)).first
                await loc.wait_for(state="visible", timeout=500)
                return
            except Exception as e:
                last_err = e
                continue
        # Fallback: textarea visible means we can type, period.
        try:
            textarea = chat_frame.locator("textarea").first
            await textarea.wait_for(state="visible", timeout=500)
            return
        except Exception as e:
            last_err = e
        await asyncio.sleep(0.4)
    raise RuntimeError(
        f"Chat never appeared ready (no greeting pattern matched and textarea "
        f"never became visible within {timeout_ms}ms): {last_err}"
    )

async def find_textarea(chat_frame, timeout_ms: int = 15000):
    candidates = [
        lambda f: f.get_by_placeholder("Type your message..").first,
        lambda f: f.locator('textarea[placeholder*="Type your message" i]').first,
        lambda f: f.locator("textarea.custom-textarea").first,
        lambda f: f.locator("textarea").first,
    ]
    deadline = time.monotonic() + timeout_ms / 1000
    last_err: Optional[Exception] = None
    while time.monotonic() < deadline:
        for build in candidates:
            try:
                loc = build(chat_frame)
                await loc.wait_for(state="visible", timeout=1500)
                return loc
            except Exception as e:
                last_err = e
                continue
        await asyncio.sleep(0.3)
    raise RuntimeError(f"Chat textarea not found: {last_err}")

async def wait_for_login_success(chat_frame, timeout_ms: int = 30000) -> None:
    """
    After the OTP is verified on the login popup, the chat widget shows a
    "Successfully signed in!" separator. We wait for it as a positive
    confirmation that login completed and the widget is in its post-login
    state. Raises RuntimeError on timeout — the worker's INNER_RETRIES loop
    catches it and re-runs the whole flow from sending the query.

    Selector tries text-based first (most stable across DOM changes), then
    the .end-separator-wrapper class with text constraint as a fallback.
    """
    deadline = time.monotonic() + timeout_ms / 1000
    last_err: Optional[Exception] = None
    candidates = [
        lambda f: f.get_by_text("Successfully signed in!").first,
        lambda f: f.get_by_text(re.compile(r"(?i)successfully\s+signed\s+in")).first,
        lambda f: f.locator(
            "div.end-separator-wrapper",
            has_text=re.compile(r"(?i)successfully\s+signed\s+in"),
        ).first,
    ]
    while time.monotonic() < deadline:
        for build in candidates:
            try:
                loc = build(chat_frame)
                await loc.wait_for(state="visible", timeout=1500)
                return
            except Exception as e:
                last_err = e
                continue
        await asyncio.sleep(0.4)
    raise RuntimeError(
        f"'Successfully signed in!' confirmation never appeared in chat "
        f"widget after Verify click (timeout {timeout_ms}ms): {last_err}"
    )


def _login_target_candidates():
    """
    Return the ordered list of locator-builders that wait_for_login_target /
    poll_for_login_target use. Defined as a function (not a module constant)
    because the regexes need fresh re.compile per call to keep them
    re-entrant — callers iterate this multiple times.
    """
    login_re_exact = re.compile(r"(?i)^\s*log[\s\-]*in\s*$")
    login_re_sub = re.compile(r"(?i)\blog[\s\-]*in\b")
    sign_in_re = re.compile(r"(?i)sign\s*in")
    sign_in_atb_re = re.compile(r"(?i)sign\s*in.*atb")

    return [
        # === LINK-STYLE — most specific first ===
        # Anchors pointing to ATB's login domains.
        lambda f: f.locator("a[href*='login-uat.atb.com']").first,
        lambda f: f.locator("a[href*='login.atb.com']").first,
        lambda f: f.locator("a[href*='atb.com']").first,
        # Anchors with combined "sign in" + "ATB" text (e.g. "Sign in to ATB Personal").
        lambda f: f.locator("a", has_text=sign_in_atb_re).first,
        lambda f: f.locator("a", has_text=re.compile(r"(?i)atb\s*(?:login|sign)")).first,
        # Generic anchors with sign-in / log-in text.
        lambda f: f.locator("a", has_text=sign_in_re).first,
        lambda f: f.locator("a", has_text=login_re_sub).first,
        # Original DEV bot-agent-response wrappers.
        lambda f: f.locator("span.bot-agent-response p a").first,
        lambda f: f.locator(".bot-agent-response p a").first,
        lambda f: f.locator(".bot-agent-response a").first,

        # === CARD-STYLE — "Log In" wording ===
        lambda f: f.locator('button.tab-focused-btn', has_text=login_re_sub).first,
        lambda f: f.locator('button', has_text=login_re_sub).first,
        lambda f: f.locator('div.title', has_text="Login").first,
        lambda f: f.locator('div.title', has_text="Log In").first,
        lambda f: f.locator('div.title').filter(has_text=login_re_exact).first,
        lambda f: f.locator('.card-button, .bot-response-card, .card')
                   .filter(has_text=login_re_sub).first,
        lambda f: f.get_by_role("button", name=login_re_exact).first,

        # === CARD-STYLE — "Sign In" wording ===
        # Same card patterns but matching the alternative button/card label
        # the bot might use on either environment.
        lambda f: f.locator('button.tab-focused-btn', has_text=sign_in_re).first,
        lambda f: f.locator('button', has_text=sign_in_re).first,
        lambda f: f.locator('div.title', has_text=sign_in_re).first,
        lambda f: f.locator('.card-button, .bot-response-card, .card')
                   .filter(has_text=sign_in_re).first,
        lambda f: f.get_by_role("button", name=sign_in_re).first,

        # === ROLE-BASED FALLBACKS ===
        lambda f: f.get_by_role("link", name=sign_in_atb_re).first,
        lambda f: f.get_by_role("link", name=re.compile(r"(?i)atb\s*login")).first,
        lambda f: f.get_by_role("link", name=sign_in_re).first,
    ]


async def _quick_check_for_login_target(chat_frame, probe_timeout_ms: int = 250):
    """
    Single fast sweep through every login-target candidate. Returns the
    first matching locator, or None if nothing matches within the budget.
    Uses a short per-candidate probe so the entire sweep finishes in well
    under a second when no element is present yet.
    """
    for build in _login_target_candidates():
        try:
            loc = build(chat_frame)
            await loc.wait_for(state="visible", timeout=probe_timeout_ms)
            return loc
        except Exception:
            continue
    return None


async def wait_for_login_target(
    chat_frame,
    timeout_ms: int = 60000,
    poll_interval_s: float = 2.0,
):
    """
    Continuously poll the chat widget for ANY clickable login element.
    Each sweep checks every candidate selector with a short probe; between
    sweeps the loop sleeps poll_interval_s. As soon as a candidate matches
    we return its locator.

    Polls roughly every poll_interval_s seconds end-to-end. So with the
    default 2s, a login element that becomes visible at any moment is
    clicked within ~2-3s, far faster than the original 30+ second effective
    cadence.
    """
    deadline = time.monotonic() + timeout_ms / 1000
    while time.monotonic() < deadline:
        loc = await _quick_check_for_login_target(chat_frame)
        if loc is not None:
            return loc
        # Don't oversleep — if there's <poll_interval left, sleep just to
        # the deadline so the final check fires close to the boundary.
        remaining = deadline - time.monotonic()
        if remaining <= 0:
            break
        await asyncio.sleep(min(poll_interval_s, remaining))
    # One last quick check before raising — the element may have appeared
    # during the final sleep tick.
    loc = await _quick_check_for_login_target(chat_frame)
    if loc is not None:
        return loc
    # Timeout — log a snippet of the chat body so the user can see what
    # text was actually there and extend the selectors if needed.
    try:
        tail = (await chat_frame.locator("body").inner_text(timeout=3000))[-600:]
        tail = tail.replace("\n", " | ").strip()
        print(f"  [diag] chat frame tail: {tail!r}", file=sys.stderr)
    except Exception:
        pass
    raise RuntimeError(
        f"Login target (card or link) never appeared in chat widget after "
        f"polling every {poll_interval_s}s for {timeout_ms / 1000:.0f}s"
    )


async def wait_for_login_link(chat_frame, timeout_ms: int = 60000):
    """
    DEV bot login: the bot replies with a paragraph inside .bot-agent-response
    that contains an <a> link (typically text "ATB login") pointing at
    login-uat.atb.com. We click whatever <a> is inside that response,
    regardless of its text content, so this is robust to wording changes.
    """
    candidates = [
        # Most specific — the link inside the response container.
        lambda f: f.locator("span.bot-agent-response p a").first,
        lambda f: f.locator(".bot-agent-response p a").first,
        lambda f: f.locator(".bot-agent-response a").first,
        # Text-based fallbacks if the class structure ever shifts.
        lambda f: f.get_by_role("link", name=re.compile(r"(?i)atb\s*login")).first,
        lambda f: f.locator("a", has_text=re.compile(r"(?i)atb\s*login")).first,
    ]
    deadline = time.monotonic() + timeout_ms / 1000
    last_err: Optional[Exception] = None
    while time.monotonic() < deadline:
        for build in candidates:
            try:
                loc = build(chat_frame)
                await loc.wait_for(state="visible", timeout=1500)
                return loc
            except Exception as e:
                last_err = e
                continue
        await asyncio.sleep(0.4)
    try:
        tail = (await chat_frame.locator("body").inner_text(timeout=3000))[-600:]
        tail = tail.replace("\n", " | ").strip()
        print(f"  [diag] chat frame tail: {tail!r}", file=sys.stderr)
    except Exception:
        pass
    raise RuntimeError(
        f"DEV login link inside .bot-agent-response never appeared: {last_err}"
    )


async def wait_for_login_card(chat_frame, timeout_ms: int = 60000):
    login_re_exact = re.compile(r"(?i)^\s*log[\s\-]*in\s*$")
    login_re_sub = re.compile(r"(?i)\blog[\s\-]*in\b")
    candidates = [
        lambda f: f.locator('button.tab-focused-btn', has_text=login_re_sub).first,
        lambda f: f.locator('button', has_text=login_re_sub).first,
        lambda f: f.locator('div.title', has_text="Login").first,
        lambda f: f.locator('div.title', has_text="Log In").first,
        lambda f: f.locator('div.title').filter(has_text=login_re_exact).first,
        lambda f: f.locator('.card-button, .bot-response-card, .card')
                   .filter(has_text=login_re_sub).first,
        lambda f: f.get_by_role("button", name=login_re_exact).first,
    ]
    deadline = time.monotonic() + timeout_ms / 1000
    last_err: Optional[Exception] = None
    while time.monotonic() < deadline:
        for build in candidates:
            try:
                loc = build(chat_frame)
                await loc.wait_for(state="visible", timeout=1500)
                return loc
            except Exception as e:
                last_err = e
                continue
        await asyncio.sleep(0.4)
    try:
        tail = (await chat_frame.locator("body").inner_text(timeout=3000))[-600:]
        tail = tail.replace("\n", " | ").strip()
        print(f"  [diag] chat frame tail: {tail!r}", file=sys.stderr)
    except Exception:
        pass
    raise RuntimeError(f"'Log In' response card never appeared: {last_err}")

# ---------- OTP via mailinator ----------
def _parse_received_to_minutes(text: str) -> Optional[int]:
    t = text.strip().lower()
    if not t:
        return None
    if t in ("just now", "now"):
        return 0
    m = re.match(r"(\d+)\s*(sec|second|seconds|min|minute|minutes|hr|hour|hours|day|days)", t)
    if not m:
        return None
    n = int(m.group(1))
    unit = m.group(2)
    if unit.startswith("sec"):
        return 0
    if unit.startswith("min"):
        return n
    if unit.startswith("hr") or unit.startswith("hour"):
        return n * 60
    if unit.startswith("day"):
        return n * 1440
    return None

async def fetch_otp_from_mailinator(
    context: BrowserContext,
    login_id: str,
    freshness_minutes: int = OTP_FRESHNESS_MINUTES,
    timeout_s: int = OTP_FETCH_TIMEOUT_S,
    max_attempts: int = OTP_MAX_ATTEMPTS,
) -> str:
    url = MAILINATOR_URL_TEMPLATE.format(login_id=login_id)
    mail_page = await context.new_page()
    deadline = time.monotonic() + timeout_s
    last_err: str = "no attempts made"
    attempts = 0  # counts each mailinator load/reload, including the initial one
    try:
        first_load = True
        while time.monotonic() < deadline:
            if attempts >= max_attempts:
                last_err = (
                    f"reached OTP_MAX_ATTEMPTS={max_attempts} "
                    f"(last: {last_err})"
                )
                break
            attempts += 1
            try:
                if first_load:
                    await mail_page.goto(url, wait_until="domcontentloaded", timeout=30000)
                    first_load = False
                else:
                    await mail_page.reload(wait_until="domcontentloaded", timeout=30000)
            except Exception as e:
                last_err = f"navigation failed: {e}"
                await asyncio.sleep(2)
                continue
            await asyncio.sleep(1.5)
            first_row = mail_page.locator("table.jambo_table tbody tr").first
            try:
                await first_row.wait_for(state="visible", timeout=5000)
            except Exception:
                last_err = "no email rows yet"
                print(
                    f"  [mailinator] no emails yet for {login_id}; "
                    f"waiting {OTP_POLL_INTERVAL_S}s before refresh"
                )
                await asyncio.sleep(OTP_POLL_INTERVAL_S)
                continue
            cells = first_row.locator("td")
            try:
                subject_text = (await cells.nth(2).inner_text()).strip()
                received_text = (await cells.nth(3).inner_text()).strip()
            except Exception as e:
                last_err = f"could not read cells: {e}"
                await asyncio.sleep(2)
                continue
            subj_match = re.search(r"\b(\d{6})\b", subject_text)
            if not subj_match:
                last_err = f"no 6-digit OTP in subject '{subject_text}'"
                await asyncio.sleep(2)
                continue
            otp = subj_match.group(1)
            if OTP_REQUIRE_JUST_NOW:
                # Strict mode: the received cell must literally contain
                # "just now". Anything else (even "5 seconds ago") is
                # rejected, and we wait for the next refresh.
                if "just now" in received_text.lower():
                    return otp
                last_err = (
                    f"newest OTP received='{received_text}'; waiting for "
                    f"'just now' (OTP_REQUIRE_JUST_NOW=True)"
                )
                print(
                    f"  [mailinator] newest OTP for {login_id} received="
                    f"'{received_text}'; waiting for an email marked 'just "
                    f"now' before using it. Refreshing in {OTP_POLL_INTERVAL_S}s.",
                    file=sys.stderr,
                )
                await asyncio.sleep(OTP_POLL_INTERVAL_S)
                continue
            minutes_ago = _parse_received_to_minutes(received_text)
            if minutes_ago is None:
                last_err = f"couldn't parse received time '{received_text}'"
                await asyncio.sleep(2)
                continue
            if minutes_ago <= freshness_minutes:
                return otp
            last_err = (
                f"newest OTP is {minutes_ago} min old "
                f"(threshold {freshness_minutes}); waiting for fresher email"
            )
            print(
                f"  [mailinator] newest OTP for {login_id} is {minutes_ago} min old "
                f"(>{freshness_minutes}); waiting {OTP_POLL_INTERVAL_S}s before refresh"
            )
            await asyncio.sleep(OTP_POLL_INTERVAL_S)
        raise RuntimeError(
            f"mailinator: could not get a fresh OTP after {attempts} attempt(s) "
            f"— last: {last_err}"
        )
    finally:
        try:
            await mail_page.close()
        except Exception:
            pass

# ---------- core: full per-query flow ----------
async def process_query(
    browser: Browser,
    query: str,
    login_id: str,
    password: str,
    screenshot_label: Optional[str] = None,
) -> tuple[str, str]:
    """
    Run the full UI flow for one row and return (conversation_id, user_id).

    screenshot_label: optional string included in the failure-screenshot
                      filename (e.g. f"row{idx+2}_attempt{n}") so multiple
                      failed attempts on the same row produce distinct files.
    """
    if not login_id or not password:
        raise RuntimeError(
            "LoginID or Password missing in the workbook for this row — can't run login flow"
        )
    context = await new_context(browser)
    page = await context.new_page()

    conv_future: asyncio.Future = asyncio.get_running_loop().create_future()
    captured: dict = {"conv_id": None, "user_id": None}

    def _on_request(req: Request) -> None:
        try:
            if req.method != "POST" or WEBHOOK_URL_SUBSTR not in req.url:
                return
            body: dict = {}
            try:
                body = req.post_data_json or {}
            except Exception:
                body = {}
            if not isinstance(body, dict):
                body = {}
            if not body:
                raw = req.post_data
                if raw:
                    try:
                        body = json.loads(raw)
                    except Exception:
                        body = {}
            request_body = (body.get("requestBody") or {}) if isinstance(body, dict) else {}
            if not isinstance(request_body, dict):
                request_body = {}
            cid = request_body.get("conversationId")
            user_details = request_body.get("userDetails") or {}
            if not isinstance(user_details, dict):
                user_details = {}
            uid = user_details.get("userId")
            if cid and not captured["conv_id"]:
                captured["conv_id"] = cid
            if uid and not captured["user_id"]:
                captured["user_id"] = uid
            if captured["conv_id"] and captured["user_id"] and not conv_future.done():
                conv_future.set_result((captured["conv_id"], captured["user_id"]))
        except Exception:
            pass

    context.on("request", _on_request)

    try:
        await page.goto(
            DEMO_URL, wait_until="domcontentloaded", timeout=ATTEMPT_TIMEOUT_S * 1000
        )
        await _rand_sleep(0.3, 0.8)

        # Open the chat widget. Behavior depends on the resolved Let's-talk
        # mode (per-environment by default, or the user's explicit override).
        effective_lets_talk = _resolve_lets_talk_mode()
        print(
            f"  [lets_talk] mode={LETS_TALK_MODE} effective={effective_lets_talk} "
            f"(env={BOT_ENVIRONMENT})",
            file=sys.stderr,
        )
        if effective_lets_talk == "ALWAYS":
            # Click first, then wait for iframe. If find_lets_talk genuinely
            # can't locate the trigger we still fall through to the iframe-
            # attach wait, but log loudly so the user can tell whether the
            # click happened or got skipped.
            try:
                lets_talk = await find_lets_talk(page, timeout_ms=30000)
                if await _try_click_cascade(lets_talk):
                    print(
                        "  [lets_talk] ALWAYS: clicked Let's talk! successfully.",
                        file=sys.stderr,
                    )
                else:
                    print(
                        "[WARN] Found Let's talk! trigger but all six click "
                        "strategies failed; relying on iframe-attach wait.",
                        file=sys.stderr,
                    )
                await _rand_sleep(0.4, 0.9)
            except Exception as e:
                print(
                    f"[WARN] [lets_talk] ALWAYS: find_lets_talk failed within "
                    f"30s — no click happened. Will still wait for iframe in "
                    f"case the chat opened on its own. Error: {e}",
                    file=sys.stderr,
                )
            await page.wait_for_selector(
                f"#{CHAT_IFRAME_ID}", state="attached", timeout=30000
            )
        elif effective_lets_talk == "NEVER":
            # No click. Just wait for the chat iframe to attach on its own.
            await page.wait_for_selector(
                f"#{CHAT_IFRAME_ID}", state="attached", timeout=30000
            )
        else:  # "AUTO"
            # Use "textarea visible in any frame" as the readiness signal,
            # not iframe-attached. Some deployments (DEV environments) pre-
            # attach #netomiChatWindow on page load even when only the Let's
            # talk! trigger is showing, which makes iframe-attached a useless
            # signal. Textarea-visible reliably means we can type.
            chat_open = False
            poll_deadline = time.monotonic() + 4.0  # 4s budget for auto-open
            while time.monotonic() < poll_deadline:
                for frame in page.frames:
                    try:
                        if await frame.locator("textarea").first.is_visible():
                            chat_open = True
                            break
                    except Exception:
                        continue
                if chat_open:
                    break
                await asyncio.sleep(0.3)
            if not chat_open:
                try:
                    lets_talk = await find_lets_talk(page, timeout_ms=30000)
                    if await _try_click_cascade(lets_talk):
                        print(
                            "  [lets_talk] AUTO: textarea wasn't visible; "
                            "clicked Let's talk! successfully.",
                            file=sys.stderr,
                        )
                    else:
                        print(
                            "[WARN] Found Let's talk! trigger but all six "
                            "click strategies failed; relying on iframe-"
                            "attach wait.",
                            file=sys.stderr,
                        )
                    await _rand_sleep(0.4, 0.9)
                except Exception as e:
                    print(
                        f"[WARN] [lets_talk] AUTO: find_lets_talk failed — "
                        f"no click. Will still wait for iframe. Error: {e}",
                        file=sys.stderr,
                    )
            else:
                print(
                    "  [lets_talk] AUTO: textarea already visible; skipping "
                    "click (chat appears auto-opened).",
                    file=sys.stderr,
                )
            await page.wait_for_selector(
                f"#{CHAT_IFRAME_ID}", state="attached", timeout=30000
            )

        chat_frame = page.frame_locator(f"#{CHAT_IFRAME_ID}")

        if await click_agree(page, chat_frame, timeout_ms=20000):
            await _rand_sleep(0.4, 0.9)

        await wait_for_greeting(chat_frame, timeout_ms=30000)
        await _rand_sleep(0.3, 0.7)

        # Send the query and continuously poll for the bot's reply (login
        # card or link). The wait_for_login_target loop polls every
        # LOGIN_LINK_POLL_INTERVAL_S seconds, so a slow bot reply is caught
        # within ~2-3s of appearing. If the per-attempt timeout elapses
        # without a hit, we re-send the query IMMEDIATELY (no extra sleep)
        # and keep polling — the polling loop itself is the only wait.
        login_target = None
        last_login_err: Optional[Exception] = None
        total_query_attempts = LOGIN_LINK_QUERY_RETRIES + 1
        for query_attempt in range(1, total_query_attempts + 1):
            textarea = await find_textarea(chat_frame, timeout_ms=15000)
            await human_type(textarea, query)
            await textarea.press("Enter")
            if query_attempt > 1:
                print(
                    f"  [login_wait] re-sent query "
                    f"(attempt {query_attempt}/{total_query_attempts})",
                    file=sys.stderr,
                )
            else:
                print(
                    f"  [login_wait] sent query; polling every "
                    f"{LOGIN_LINK_POLL_INTERVAL_S}s for the login element.",
                    file=sys.stderr,
                )

            wait_ms = LOGIN_LINK_WAIT_TIMEOUT_S * 1000
            try:
                # Unified waiter — handles both card-style and link-style
                # bot replies, regardless of BOT_ENVIRONMENT. Polls every
                # LOGIN_LINK_POLL_INTERVAL_S so a late bot response is
                # picked up quickly without waiting out the full window.
                login_target = await wait_for_login_target(
                    chat_frame,
                    timeout_ms=wait_ms,
                    poll_interval_s=LOGIN_LINK_POLL_INTERVAL_S,
                )
                break  # Got it — exit the retry loop.
            except Exception as e:
                last_login_err = e
                if query_attempt < total_query_attempts:
                    print(
                        f"[WARN] [login_wait] login element didn't appear "
                        f"within {LOGIN_LINK_WAIT_TIMEOUT_S}s on attempt "
                        f"{query_attempt}/{total_query_attempts}; re-sending "
                        f"the query immediately and continuing to poll. "
                        f"Error: {e}",
                        file=sys.stderr,
                    )
                    # No sleep here — the polling resumes inside the next
                    # wait_for_login_target call right away, so a bot reply
                    # that lands during the type-and-press window still gets
                    # caught within poll_interval_s.

        if login_target is None:
            raise RuntimeError(
                f"Login link/card never appeared after {total_query_attempts} "
                f"query attempts (each waiting {LOGIN_LINK_WAIT_TIMEOUT_S}s). "
                f"Last error: {last_login_err}"
            )

        async with context.expect_page(timeout=15000) as popup_info:
            await human_click(login_target)
        login_page: Page = await popup_info.value
        await login_page.wait_for_load_state(
            "domcontentloaded", timeout=ATTEMPT_TIMEOUT_S * 1000
        )
        await _rand_sleep(0.4, 1.0)

        username_input = login_page.locator("#username")
        await username_input.wait_for(state="visible", timeout=30000)
        await human_type(username_input, login_id)
        continue_btn = login_page.locator("#main-button")
        await continue_btn.wait_for(state="visible", timeout=15000)
        await human_click(continue_btn)
        await _rand_sleep(0.4, 0.9)

        password_input = login_page.locator("#password")
        await password_input.wait_for(state="visible", timeout=30000)
        await human_type(password_input, password)
        sign_on_btn = login_page.locator("#signOnButton")
        await sign_on_btn.wait_for(state="visible", timeout=15000)
        await human_click(sign_on_btn)
        await _rand_sleep(0.5, 1.0)

        passcode_input = login_page.locator("#passcode")
        await passcode_input.wait_for(state="visible", timeout=60000)

        otp = await fetch_otp_from_mailinator(
            context,
            login_id=login_id,
            freshness_minutes=OTP_FRESHNESS_MINUTES,
            timeout_s=OTP_FETCH_TIMEOUT_S,
            max_attempts=OTP_MAX_ATTEMPTS,
        )

        await login_page.bring_to_front()
        await passcode_input.wait_for(state="visible", timeout=15000)
        await passcode_input.click()
        await human_type(passcode_input, otp)
        verify_btn = login_page.locator("#sign-on")
        await verify_btn.wait_for(state="visible", timeout=15000)
        await human_click(verify_btn)
        await _rand_sleep(0.8, 1.5)

        try:
            await page.bring_to_front()
        except Exception:
            pass

        # Confirm the widget actually transitioned to the post-login state.
        # If this raises (the success banner never appears), the worker's
        # INNER_RETRIES loop catches it and re-runs the whole flow from the
        # query-send step.
        await wait_for_login_success(chat_frame, timeout_ms=30000)
        await _rand_sleep(0.3, 0.7)

        try:
            conv_id, user_id = await asyncio.wait_for(conv_future, timeout=CONV_ID_WAIT_S)
        except asyncio.TimeoutError:
            have_conv = captured.get("conv_id") or "<missing>"
            have_user = captured.get("user_id") or "<missing>"
            raise RuntimeError(
                f"Did not capture both IDs within {CONV_ID_WAIT_S}s. "
                f"conv_id={have_conv} user_id={have_user}"
            )
        if not conv_id:
            raise RuntimeError("Captured conversationId was empty")
        if not user_id:
            raise RuntimeError("Captured userId was empty")
        return conv_id, user_id

    except Exception:
        # Snapshot the page right at the failure point so the user can see
        # what state the widget was in (modal still up? login form blank?
        # OTP page stuck loading? etc.). Best-effort — don't let screenshot
        # errors mask the real exception.
        if UI_SCREENSHOT_ON_FAILURE:
            try:
                screenshot_dir = Path(UI_SCREENSHOT_DIR).resolve()
                screenshot_dir.mkdir(parents=True, exist_ok=True)
                ts = time.strftime("%Y%m%d_%H%M%S", time.localtime())
                ts += f"_{time.time_ns() % 1_000_000:06d}"
                safe_login = re.sub(r"[^a-zA-Z0-9_-]", "_", login_id or "unknown")
                label = (
                    re.sub(r"[^a-zA-Z0-9_-]", "_", screenshot_label)
                    if screenshot_label else "fail"
                )
                fname = f"{label}_{safe_login}_{ts}.png"
                path = screenshot_dir / fname
                await page.screenshot(path=str(path), full_page=True)
                print(
                    f"[DIAG] saved failure screenshot → {path}",
                    file=sys.stderr,
                )
            except Exception as snap_err:
                print(
                    f"[DIAG] couldn't take failure screenshot: {snap_err}",
                    file=sys.stderr,
                )
        raise  # re-raise the original exception so the worker handles it

    finally:
        try:
            context.remove_listener("request", _on_request)
        except Exception:
            pass
        try:
            await context.close()
        except Exception:
            pass

# ---------- worker orchestration ----------
async def worker(
    sem: asyncio.Semaphore,
    browser: Browser,
    row_index: int,
    query: str,
    login_id: str,
    password: str,
    rows: list[list[str]],
    lock: asyncio.Lock,
    xlsx_path: Path,
    header: list[str],
) -> None:
    async with sem:
        last_err: Optional[Exception] = None
        for attempt in range(1, INNER_RETRIES + 1):
            try:
                conv_id, user_id = await process_query(
                    browser, query, login_id, password,
                    screenshot_label=f"row{row_index + 2}_attempt{attempt}",
                )
                async with lock:
                    while len(rows[row_index]) < NUM_COLUMNS:
                        rows[row_index].append("")
                    rows[row_index][COL_CONV_ID] = conv_id
                    rows[row_index][COL_USER_ID] = user_id or ""
                    save_xlsx(xlsx_path, header, rows)
                print(
                    f"[OK]   row {row_index + 2}: {query[:60]!r} -> "
                    f"conv={conv_id} user={user_id or '<missing>'}"
                )
                return
            except Exception as e:
                last_err = e
                print(
                    f"[WARN] row {row_index + 2} attempt {attempt}/{INNER_RETRIES} failed: {e}",
                    file=sys.stderr,
                )
                await asyncio.sleep(RETRY_BACKOFF_BASE_S * attempt)
        # All retries exhausted — write the error to the row so the API
        # phase will skip it cleanly.
        async with lock:
            while len(rows[row_index]) < NUM_COLUMNS:
                rows[row_index].append("")
            rows[row_index][COL_CONV_ID] = f"ERROR: {last_err}"[:200]
            rows[row_index][COL_USER_ID] = ""
            save_xlsx(xlsx_path, header, rows)
        print(
            f"[FAIL] row {row_index + 2}: {query[:60]!r} -> {last_err}", file=sys.stderr
        )

async def run_ui_flow(workers: int) -> int:
    xlsx_path = Path(INPUT_XLSX).resolve()
    if not xlsx_path.exists():
        print(f"Input workbook not found: {xlsx_path}", file=sys.stderr)
        return 2

    if UI_PHASE_MODE == "REFRESH_ALL":
        rows, header = load_xlsx(xlsx_path)
        header = normalize_header(header)
        wiped = 0
        for r in rows:
            while len(r) < NUM_COLUMNS:
                r.append("")
            if (r[COL_CONV_ID] or "").strip() or (r[COL_USER_ID] or "").strip():
                wiped += 1
            r[COL_CONV_ID] = ""
            r[COL_USER_ID] = ""
        save_xlsx(xlsx_path, header, rows)
        print(
            f"UI_PHASE_MODE=REFRESH_ALL — cleared conversationId/userId on {wiped} row(s) before start."
        )
    else:
        # FILL_MISSING: only rows with blank or ERROR D/E will be picked up by needs_work().
        print("UI_PHASE_MODE=FILL_MISSING — leaving existing conversationId/userId in place.")

    async with async_playwright() as pw:
        browser = await launch_browser(pw)
        try:
            pass_num = 0
            while pass_num < MAX_PASSES:
                pass_num += 1
                rows, header = load_xlsx(xlsx_path)
                header = normalize_header(header)
                todo = [
                    (i, r[COL_QUERY], r[COL_LOGIN_ID], r[COL_PASSWORD])
                    for i, r in enumerate(rows)
                    if r and r[COL_QUERY].strip() and needs_work(r)
                ]
                if not todo:
                    print(f"All {len(rows)} rows have a conversationId. UI flow done.")
                    return 0
                print(
                    f"=== UI Pass {pass_num}: {len(todo)} rows remaining, "
                    f"browser={BROWSER}, workers={workers}, "
                    f"headless={HEADLESS}, human_like={HUMAN_LIKE} ==="
                )
                sem = asyncio.Semaphore(workers)
                lock = asyncio.Lock()
                tasks = [
                    asyncio.create_task(worker(
                        sem, browser, idx, q, lid, pw_, rows, lock, xlsx_path, header,
                    ))
                    for (idx, q, lid, pw_) in todo
                ]
                await asyncio.gather(*tasks, return_exceptions=True)
                await asyncio.sleep(INTER_PASS_DELAY_S)
            rows, _ = load_xlsx(xlsx_path)
            remaining = [
                (i + 2, r[COL_QUERY]) for i, r in enumerate(rows) if needs_work(r)
            ]
            if remaining:
                print(
                    f"Hit MAX_PASSES={MAX_PASSES}. {len(remaining)} row(s) still unfilled:",
                    file=sys.stderr,
                )
                for line, q in remaining:
                    print(f"  row {line}: {q!r}", file=sys.stderr)
                return 1
            return 0
        finally:
            await browser.close()

def _import_api_runner() -> Optional[object]:
    """Try both filename casings for the API runner sibling module."""
    last_err: Optional[Exception] = None
    for candidate in ("atb_api_runner", "ATB_API_Runner"):
        try:
            return __import__(candidate)
        except ImportError as e:
            last_err = e
    print(
        f"[ERROR] Could not import the API runner. Expected a sibling file "
        f"named atb_api_runner.py or ATB_API_Runner.py in the same directory. "
        f"Last import error: {last_err}",
        file=sys.stderr,
    )
    return None


async def run_pipelined_flow(workers: int) -> int:
    """
    Pipelined orchestrator with concurrent UI/API stages.

    Each iteration:
      1. Kick off the UI batch (UI_WORKERS rows) as a task.
      2. Concurrently await BOTH that UI task AND the previous batch's API
         task — so UI batch K+1 runs in parallel with API batch K.
      3. When both complete, kick off API batch K+1 as a new pending task and
         loop. The API batch uses skip_done=True so it only processes TestIDs
         whose users just became available.

    UI batches never overlap with each other (one browser-batch at a time).
    API batches never overlap with each other either (the previous one is
    awaited before the next one is dispatched). The overlap is strictly
    UI-batch-K+1 vs API-batch-K, which doesn't share write targets — UI
    writes to queries.xlsx, API writes to the active test sheet (per env).
    """
    queries_path = Path(INPUT_XLSX).resolve()
    if not queries_path.exists():
        print(f"Input workbook not found: {queries_path}", file=sys.stderr)
        return 2

    test_sheet_path = Path(TEST_SHEET_XLSX).resolve()
    if not test_sheet_path.exists():
        print(f"Test sheet workbook not found: {test_sheet_path}", file=sys.stderr)
        return 2

    api_runner = _import_api_runner()
    if api_runner is None:
        return 3

    if UI_PHASE_MODE == "REFRESH_ALL":
        rows, header = load_xlsx(queries_path)
        header = normalize_header(header)
        wiped = 0
        for r in rows:
            while len(r) < NUM_COLUMNS:
                r.append("")
            if (r[COL_CONV_ID] or "").strip() or (r[COL_USER_ID] or "").strip():
                wiped += 1
            r[COL_CONV_ID] = ""
            r[COL_USER_ID] = ""
        save_xlsx(queries_path, header, rows)
        print(
            f"UI_PHASE_MODE=REFRESH_ALL — cleared conversationId/userId on "
            f"{wiped} row(s) before pipelined start."
        )
    else:
        print(
            "UI_PHASE_MODE=FILL_MISSING — leaving existing conversationId/userId "
            "in place for pipelined start."
        )

    # Per-batch API stats so the final summary tells the user the truth.
    api_batch_stats = {"runs": 0, "rc_zero": 0, "rc_nonzero": 0, "crashed": 0}

    async def _run_api_batch() -> None:
        api_batch_stats["runs"] += 1
        try:
            rc = await api_runner.run_api_flow(
                queries_path=queries_path,
                test_sheet_path=test_sheet_path,
                workers=workers,
                manual_assignment=MANUAL_ASSIGNMENT,
                turn_retries=API_TURN_RETRIES,
                turn_timeout_s=API_TURN_TIMEOUT_S,
                inter_turn_delay_s=API_INTER_TURN_DELAY_S,
                get_max_attempts=API_GET_MAX_ATTEMPTS,
                get_retry_delay_s=API_GET_RETRY_DELAY_S,
                skip_done=True,
                bot_ref_id=_get_active_bot_ref_id(),
            )
            if rc == 0:
                api_batch_stats["rc_zero"] += 1
            else:
                api_batch_stats["rc_nonzero"] += 1
                # rc=1 commonly means "overflow TestIDs" in pipelined mode,
                # which is normal until later batches catch up. We log without
                # making it sound alarming.
                print(
                    f"[INFO] API batch returned rc={rc} "
                    f"(non-zero is expected mid-pipeline if some TestIDs "
                    f"don't have users captured yet)",
                    file=sys.stderr,
                )
        except Exception as e:
            api_batch_stats["crashed"] += 1
            print(f"[ERROR] API batch crashed: {e}", file=sys.stderr)

    # Note on pipelined timing semantics: the API task we dispatch reads
    # queries.xlsx fresh at *its* start. Depending on event-loop scheduling
    # vs. how fast the next UI batch's workers save, that load may happen
    # before or after the next batch's writes. So a given API batch can pick
    # up MORE TestIDs than just the row from "its" UI batch — it processes
    # whatever undone TestIDs have available users at the moment it loads.
    # That's by design and 1:1 user-to-TestID still holds; the tradeoff is
    # that batch boundaries in the logs aren't deterministic.
    pending_api_task: Optional[asyncio.Task] = None

    async with async_playwright() as pw_inst:
        browser = await launch_browser(pw_inst)
        try:
            pass_num = 0
            while pass_num < MAX_PASSES:
                pass_num += 1
                rows, header = load_xlsx(queries_path)
                header = normalize_header(header)
                todo = [
                    (i, r[COL_QUERY], r[COL_LOGIN_ID], r[COL_PASSWORD])
                    for i, r in enumerate(rows)
                    if r and r[COL_QUERY].strip() and needs_work(r)
                ]
                if not todo:
                    print(
                        f"All {len(rows)} rows have a conversationId. "
                        f"UI flow done."
                    )
                    break
                print(
                    f"=== UI Pass {pass_num} (pipelined): {len(todo)} rows "
                    f"remaining, browser={BROWSER}, workers={workers}, "
                    f"headless={HEADLESS} ==="
                )
                # Walk needs-work rows in batches of `workers` size.
                for batch_start in range(0, len(todo), workers):
                    batch = todo[batch_start:batch_start + workers]
                    batch_lines = [b[0] + 2 for b in batch]
                    print(
                        f"  --- UI batch (pass {pass_num}, sheet rows "
                        f"{batch_lines}) ---"
                    )
                    # Each UI worker is its own task (already scheduled and
                    # running). We await them all together with the previous
                    # batch's API task so the UI batch and prior API run
                    # concurrently.
                    sem = asyncio.Semaphore(workers)
                    lock = asyncio.Lock()
                    ui_worker_tasks = [
                        asyncio.create_task(worker(
                            sem, browser, idx, q, lid, pw_,
                            rows, lock, queries_path, header,
                        ))
                        for (idx, q, lid, pw_) in batch
                    ]

                    if pending_api_task is not None:
                        print(
                            f"  --- (UI batch is running concurrently with "
                            f"prior API batch) ---"
                        )
                        # gather() with return_exceptions=True won't raise
                        # if a worker or the API task errors — both already
                        # log their own failures.
                        await asyncio.gather(
                            *ui_worker_tasks, pending_api_task,
                            return_exceptions=True,
                        )
                    else:
                        await asyncio.gather(
                            *ui_worker_tasks, return_exceptions=True,
                        )

                    # UI batch done → dispatch API batch for it. Don't await:
                    # the next iteration's UI batch will run alongside it.
                    print(
                        f"  --- API batch dispatched (pass {pass_num}, "
                        f"post-rows {batch_lines}; will run in parallel with "
                        f"the next UI batch) ---"
                    )
                    pending_api_task = asyncio.create_task(_run_api_batch())

                await asyncio.sleep(INTER_PASS_DELAY_S)

            # Drain the last pending API task before the final sweep.
            if pending_api_task is not None:
                print("=== Waiting for final pending API batch to finish ===")
                try:
                    await pending_api_task
                except Exception as e:
                    print(f"[ERROR] Final API batch crashed: {e}", file=sys.stderr)
                pending_api_task = None

            # Final API sweep — catches any leftover TestIDs in case skip_done
            # missed something or earlier batches errored out.
            print("=== Final API sweep (skip_done) ===")
            await _run_api_batch()

            rows, _ = load_xlsx(queries_path)
            remaining = [
                (i + 2, r[COL_QUERY]) for i, r in enumerate(rows) if needs_work(r)
            ]

            # Pipelining summary — truthfully report what each API batch did.
            print("=" * 72)
            print("PIPELINED FLOW SUMMARY")
            print(f"  API batches dispatched:       {api_batch_stats['runs']}")
            print(f"  API batches rc=0 (clean):     {api_batch_stats['rc_zero']}")
            print(f"  API batches rc!=0 (partial):  {api_batch_stats['rc_nonzero']}")
            print(f"  API batches crashed:          {api_batch_stats['crashed']}")
            print("=" * 72)

            if remaining:
                print(
                    f"Hit MAX_PASSES={MAX_PASSES}. {len(remaining)} row(s) "
                    f"still unfilled:",
                    file=sys.stderr,
                )
                for line, q in remaining:
                    print(f"  row {line}: {q!r}", file=sys.stderr)
                return 1
            if api_batch_stats["crashed"] > 0:
                return 1
            return 0
        finally:
            await browser.close()


def _clamp_workers(n: int, label: str) -> int:
    if n < 1:
        print(f"[WARN] {label}={n} clamped to 1", file=sys.stderr)
        return 1
    if n > MAX_WORKERS:
        print(f"[WARN] {label}={n} clamped to MAX_WORKERS={MAX_WORKERS}", file=sys.stderr)
        return MAX_WORKERS
    return n

_VALID_UI_MODES = {"REFRESH_ALL", "FILL_MISSING", "SKIP"}
_VALID_LETS_TALK_MODES = {"DEFAULT", "AUTO", "ALWAYS", "NEVER"}
_VALID_BOT_ENVIRONMENTS = {"STAGING", "DEV"}

# Per-environment defaults used when LETS_TALK_MODE is "DEFAULT". STAGING's
# widget auto-opens on page load (clicking would close it), so we don't click.
# DEV's widget needs an explicit click every time. If a future environment is
# added without an entry here, the resolver falls back to "AUTO".
_LETS_TALK_DEFAULTS_BY_ENV = {
    "STAGING": "NEVER",
    "DEV": "ALWAYS",
}


def _resolve_lets_talk_mode() -> str:
    """
    Return the effective Let's-talk handling mode for the current
    BOT_ENVIRONMENT. If LETS_TALK_MODE is one of the explicit overrides
    (AUTO/ALWAYS/NEVER), use it as-is. If it's "DEFAULT", pick from the
    per-env table.
    """
    if LETS_TALK_MODE != "DEFAULT":
        return LETS_TALK_MODE
    return _LETS_TALK_DEFAULTS_BY_ENV.get(BOT_ENVIRONMENT, "AUTO")


def _get_active_bot_ref_id() -> str:
    """Return the Netomi x-bot-ref-id for the current BOT_ENVIRONMENT."""
    return STAGING_BOT_REF_ID if BOT_ENVIRONMENT == "STAGING" else DEV_BOT_REF_ID


def _set_bot_environment(env: str) -> None:
    """Switch BOT_ENVIRONMENT, DEMO_URL, and TEST_SHEET_XLSX at runtime.
    Used by the RUN_BOTH_ENVIRONMENTS path to flip from STAGING to DEV
    between runs without the user having to touch any config."""
    global BOT_ENVIRONMENT, DEMO_URL, TEST_SHEET_XLSX
    BOT_ENVIRONMENT = env
    DEMO_URL = STAGING_DEMO_URL if env == "STAGING" else DEV_DEMO_URL
    TEST_SHEET_XLSX = (
        TEST_SHEET_STAGING_XLSX if env == "STAGING" else TEST_SHEET_DEV_XLSX
    )
    print(
        f"BOT_ENVIRONMENT set → {env}, URL → {DEMO_URL}, "
        f"TEST_SHEET → {TEST_SHEET_XLSX}, "
        f"BOT_REF_ID → {_get_active_bot_ref_id()}, "
        f"LETS_TALK_MODE={LETS_TALK_MODE} (effective: {_resolve_lets_talk_mode()})",
        file=sys.stderr,
    )


def _clear_queries_user_columns(queries_str: str) -> None:
    """Wipe columns D (conversationId) and E (userId) from queries.xlsx,
    keeping A (query), B (LoginID), C (Password). Preserves header row."""
    path = Path(queries_str).resolve()
    if not path.exists():
        return
    rows, header = load_xlsx(path)
    for r in rows:
        while len(r) < NUM_COLUMNS:
            r.append("")
        r[COL_CONV_ID] = ""
        r[COL_USER_ID] = ""
    save_xlsx(path, header, rows)
    print(
        f"Cleared {path.name} columns D-E (conversationId, userId)",
        file=sys.stderr,
    )


def _maybe_retry_failed(rc: int) -> int:
    """If RETRY_FAILED_API_TURNS is set, run the retry-failed pass after the
    main flow finishes. Returns the worse of (incoming rc, retry rc)."""
    if not RETRY_FAILED_API_TURNS:
        return rc
    runner = _import_api_runner()
    if runner is None:
        return rc if rc != 0 else 3
    retry_workers = _clamp_workers(API_WORKERS, "API_WORKERS")
    try:
        retry_rc = runner.retry_failed_sync(
            test_sheet_path=TEST_SHEET_XLSX,
            workers=retry_workers,
            turn_retries=API_TURN_RETRIES,
            turn_timeout_s=API_TURN_TIMEOUT_S,
            inter_turn_delay_s=API_INTER_TURN_DELAY_S,
            get_max_attempts=API_GET_MAX_ATTEMPTS,
            get_retry_delay_s=API_GET_RETRY_DELAY_S,
            bot_ref_id=_get_active_bot_ref_id(),
        )
    except KeyboardInterrupt:
        print("Interrupted during retry-failed pass.", file=sys.stderr)
        sys.exit(130)
    except Exception as e:
        print(f"[ERROR] Retry-failed pass crashed: {e}", file=sys.stderr)
        return rc if rc != 0 else 1
    return rc if rc != 0 else retry_rc


def _run_environment_flow() -> int:
    """
    Run UI + API + retry-failed for the CURRENT BOT_ENVIRONMENT global. Used
    once per environment, called twice when RUN_BOTH_ENVIRONMENTS is set.
    Returns the final rc for this environment (0 on full success).
    """
    if PARALLEL_PROCESSING and UI_PHASE_MODE != "SKIP":
        workers = _clamp_workers(UI_WORKERS, "UI_WORKERS")
        try:
            rc = asyncio.run(run_pipelined_flow(workers))
        except KeyboardInterrupt:
            print("Interrupted during pipelined flow.", file=sys.stderr)
            sys.exit(130)
        except Exception as e:
            print(f"[ERROR] Pipelined flow crashed: {e}", file=sys.stderr)
            return 1
        return _maybe_retry_failed(rc)

    # ---- Sequential UI phase (or SKIP) ----
    if UI_PHASE_MODE == "SKIP":
        print(
            f"UI_PHASE_MODE=SKIP — not opening a browser. Using existing "
            f"conversationId/userId values in {INPUT_XLSX}."
        )
        if PARALLEL_PROCESSING:
            print(
                "Note: PARALLEL_PROCESSING=True has no effect when "
                "UI_PHASE_MODE=SKIP — there's no UI work to interleave with "
                "API. Falling back to a single API run with API_WORKERS."
            )
        ui_rc = 0
    else:
        ui_workers = _clamp_workers(UI_WORKERS, "UI_WORKERS")
        try:
            ui_rc = asyncio.run(run_ui_flow(ui_workers))
        except KeyboardInterrupt:
            print("Interrupted during UI phase.", file=sys.stderr)
            sys.exit(130)

    # ---- API phase ----
    api_workers = _clamp_workers(API_WORKERS, "API_WORKERS")
    atb_api_runner = _import_api_runner()
    if atb_api_runner is None:
        return ui_rc if ui_rc != 0 else 3
    try:
        api_rc = atb_api_runner.run_sync(
            queries_path=INPUT_XLSX,
            test_sheet_path=TEST_SHEET_XLSX,
            workers=api_workers,
            manual_assignment=MANUAL_ASSIGNMENT,
            turn_retries=API_TURN_RETRIES,
            turn_timeout_s=API_TURN_TIMEOUT_S,
            inter_turn_delay_s=API_INTER_TURN_DELAY_S,
            get_max_attempts=API_GET_MAX_ATTEMPTS,
            get_retry_delay_s=API_GET_RETRY_DELAY_S,
            bot_ref_id=_get_active_bot_ref_id(),
        )
    except KeyboardInterrupt:
        print("Interrupted during API phase.", file=sys.stderr)
        sys.exit(130)
    except Exception as e:
        print(f"[ERROR] API flow crashed: {e}", file=sys.stderr)
        return ui_rc if ui_rc != 0 else 1

    final_rc = ui_rc if ui_rc != 0 else api_rc
    return _maybe_retry_failed(final_rc)


def main() -> None:
    # ---- Validate modes ----
    if UI_PHASE_MODE not in _VALID_UI_MODES:
        print(
            f"[ERROR] UI_PHASE_MODE={UI_PHASE_MODE!r} is not valid. "
            f"Choose one of: {sorted(_VALID_UI_MODES)}.",
            file=sys.stderr,
        )
        sys.exit(2)
    if LETS_TALK_MODE not in _VALID_LETS_TALK_MODES:
        print(
            f"[ERROR] LETS_TALK_MODE={LETS_TALK_MODE!r} is not valid. "
            f"Choose one of: {sorted(_VALID_LETS_TALK_MODES)}.",
            file=sys.stderr,
        )
        sys.exit(2)
    if BOT_ENVIRONMENT not in _VALID_BOT_ENVIRONMENTS:
        print(
            f"[ERROR] BOT_ENVIRONMENT={BOT_ENVIRONMENT!r} is not valid. "
            f"Choose one of: {sorted(_VALID_BOT_ENVIRONMENTS)}.",
            file=sys.stderr,
        )
        sys.exit(2)
    # PARALLEL_PROCESSING grows user_pool incrementally as UI batches finish.
    # MANUAL_ASSIGNMENT requires every TestID's pre-filled triple to already
    # exist in user_pool at validation time. Combining them means early API
    # batches will fail validation hard for users that haven't been captured
    # yet. Refuse the combo upfront with a clear message instead of letting
    # every batch raise mid-flight.
    if PARALLEL_PROCESSING and MANUAL_ASSIGNMENT and UI_PHASE_MODE != "SKIP":
        print(
            "[ERROR] PARALLEL_PROCESSING=True is incompatible with "
            "MANUAL_ASSIGNMENT=True when the UI phase is not SKIP. Manual "
            "assignment validates every TestID's pre-filled (conv_id, "
            "user_id, LoginID) against the user pool, but in pipelined mode "
            "the pool is built incrementally — early batches would fail "
            "validation for users not yet captured. Either set "
            "PARALLEL_PROCESSING=False, or run the UI phase first (e.g. with "
            "PARALLEL_PROCESSING=False), then re-run with UI_PHASE_MODE=SKIP "
            "and MANUAL_ASSIGNMENT=True.",
            file=sys.stderr,
        )
        sys.exit(2)
    # RUN_BOTH_ENVIRONMENTS clears the test sheet's D/E/F columns between
    # environments, which would wipe any manual assignments. Refuse this
    # combination so the user doesn't lose pre-filled data unintentionally.
    if RUN_BOTH_ENVIRONMENTS and MANUAL_ASSIGNMENT:
        print(
            "[ERROR] RUN_BOTH_ENVIRONMENTS=True is incompatible with "
            "MANUAL_ASSIGNMENT=True. The cross-environment cleanup wipes the "
            "test sheet's D/E/F columns (so DEV starts with a blank slate), "
            "which would also wipe the user-prefilled triples manual mode "
            "needs. Run environments individually instead.",
            file=sys.stderr,
        )
        sys.exit(2)

    # ---- Single-environment run (default) ----
    if not RUN_BOTH_ENVIRONMENTS:
        print(
            f"BOT_ENVIRONMENT={BOT_ENVIRONMENT} → URL={DEMO_URL}, "
            f"BOT_REF_ID={_get_active_bot_ref_id()}, "
            f"LETS_TALK_MODE={LETS_TALK_MODE} "
            f"(effective: {_resolve_lets_talk_mode()})",
            file=sys.stderr,
        )
        sys.exit(_run_environment_flow())

    # ---- RUN_BOTH_ENVIRONMENTS: STAGING then DEV ----
    # The two test sheets are independent files (TEST_SHEET_STAGING_XLSX and
    # TEST_SHEET_DEV_XLSX), so no archiving or test-sheet clearing is needed.
    # We only clear queries.xlsx D/E between runs so DEV captures fresh
    # conversationId/userId for the same login users.
    print(
        "\n" + "=" * 72 + "\n"
        f"RUN_BOTH_ENVIRONMENTS=True — running STAGING (against "
        f"{TEST_SHEET_STAGING_XLSX}), then DEV (against {TEST_SHEET_DEV_XLSX}).\n"
        + "=" * 72 + "\n",
        file=sys.stderr,
    )

    # 1. STAGING run
    _set_bot_environment("STAGING")
    rc_staging = _run_environment_flow()
    print(
        "\n" + "=" * 72 + f"\nSTAGING run finished (rc={rc_staging})\n"
        + "=" * 72 + "\n",
        file=sys.stderr,
    )

    # 2. Clear queries.xlsx D/E so DEV captures fresh IDs for its UI runs.
    #    The test sheet for DEV is separate, so nothing else needs clearing.
    print("\n--- Clearing queries.xlsx D/E for DEV run ---\n", file=sys.stderr)
    _clear_queries_user_columns(INPUT_XLSX)

    # 3. DEV run
    _set_bot_environment("DEV")
    rc_dev = _run_environment_flow()
    print(
        "\n" + "=" * 72 + f"\nDEV run finished (rc={rc_dev})\n"
        + "=" * 72 + "\n",
        file=sys.stderr,
    )

    # Combined exit code: STAGING's failure surfaces over DEV's success.
    sys.exit(rc_staging if rc_staging != 0 else rc_dev)

if __name__ == "__main__":
    main()