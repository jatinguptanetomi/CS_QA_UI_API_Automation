"""
atb_api_runner.py

API-driven multi-turn conversation runner for the ATB Financial Netomi chatbot.

Consumes:
  - queries.xlsx                 : user pool (conversationId, userId, LoginID) per row
                                   as produced by netomi_conversation_extractor.py
  - ATBFinancial_Test_Sheet.xlsx : test cases grouped by TestID

Writes back into ATBFinancial_Test_Sheet.xlsx:
  - column C (Payload)    : requestID returned by the first POST, per turn
  - column D (conversationId), column E (userId), column F (LoginID)  (in sequential mode)
  - column G (Responses)  : concatenated bot reply text extracted from the GET response

Per-turn pipeline:
  POST https://api.netomi.com/v1/webhook                        -> requestID
  GET  https://api.netomi.com/v2/ceaas/messages?requestId=<id>  -> up to 15 attempts
       with alternating bare/"_res" suffix, 2s between attempts
  extract bot reply text from responses[0].attachments with
       type == "ai.msg.domain.responses.core.Text"
       attachmentResponseType == "ANSWER_AI_RESPONSE"
       text non-empty
  concatenate with "\n\n" and write row atomically

Runs TestIDs in parallel (semaphore-gated); turns within one TestID strictly sequential.
Each TestID uses exactly one user (1:1 lock). Sequential or manual assignment.

Can be called from netomi_conversation_extractor.py via run_sync(...) at the end
of its UI flow, or executed standalone (python atb_api_runner.py) to re-run only
the API phase against existing xlsx files.
"""
from __future__ import annotations

import asyncio
import html
import json
import os
import re
import sys
import tempfile
import time
import uuid
from collections import OrderedDict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import httpx
from openpyxl import Workbook, load_workbook

# ==================== DEBUG ====================
# When True, every outgoing POST and GET prints its payload (request body) and
# response body to stdout. Keep False in normal operation; flip to True only
# when you need to inspect exactly what's going over the wire.
DEBUG_DUMP_REQUESTS: bool = False

# ==================== DEFAULTS (overridable via run_sync / run_api_flow args) ====================
DEFAULT_QUERIES_XLSX = "queries.xlsx"
DEFAULT_TEST_SHEET_XLSX = "ATBFinancial_Test_Sheet.xlsx"
DEFAULT_WORKERS = 5
DEFAULT_MANUAL_ASSIGNMENT = False
DEFAULT_TURN_RETRIES = 2
DEFAULT_TURN_TIMEOUT_S = 60.0
DEFAULT_INTER_TURN_DELAY_S = 0.5
DEFAULT_GET_MAX_ATTEMPTS = 15
DEFAULT_GET_RETRY_DELAY_S = 2.0

# ==================== NETOMI API CONTRACT ====================
POST_URL = "https://api.netomi.com/v1/webhook"
GET_URL_TEMPLATE = "https://api.netomi.com/v2/ceaas/messages?requestId={request_id}"

# Browser-shaped User-Agent. The default httpx UA (`python-httpx/X.Y.Z`) looks
# like a bot to some edge/WAF configurations and can cause the server to route
# the request to an anonymous/logged-out experience even when the conversationId
# and userId are valid. Sending a browser UA avoids that filter. The string here
# intentionally matches the `current_user_agent` value inside the POST body's
# CUSTOM_ATTRIBUTES list so the two are consistent if the server ever cross-checks.
BROWSER_USER_AGENT: str = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36"
)

# Bot reference IDs are per-environment (each Netomi bot is registered
# separately). The active value gets injected into POST_HEADERS at request
# time via the bot_ref_id parameter that's plumbed through run_api_flow /
# run_retry_failed. Standalone runs of this module fall back to
# DEFAULT_BOT_REF_ID below.
STAGING_BOT_REF_ID: str = "fab252f5-dd1b-4be7-9a39-898585ddb769"
DEV_BOT_REF_ID: str = "fe3b08f5-650b-490d-988c-34d0e6ecdf65"
DEFAULT_BOT_REF_ID: str = STAGING_BOT_REF_ID

# x-bot-ref-id is intentionally NOT in this dict — it's added per-request
# inside _post_message based on the bot_ref_id parameter.
POST_HEADERS: Dict[str, str] = {
    "x_integration": "NETOMI_WEB_WIDGET",
    "TEST_REQUEST": "false",
    "Content-Type": "application/json",
    "x-channel": "CHAT",
    "User-Agent": BROWSER_USER_AGENT,
}

# Static portion of the POST body — same byte-for-byte for every turn of every TestID.
POST_BODY_STATIC_CUSTOM_ATTRIBUTES: List[Dict[str, Any]] = [
    {"type": "TEXT", "name": "widget_id", "value": None, "scope": "LIFE_TIME"},
    {"type": "TEXT", "name": "user_timezone", "value": "Asia/Calcutta", "scope": "LIFE_TIME"},
    {"type": "TEXT", "name": "visitor_url", "value": "https://demo.pomsgai.io", "scope": "LIFE_TIME"},
    {
        "type": "TEXT",
        "name": "current_user_agent",
        "value": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36"
        ),
        "scope": "LIFE_TIME",
    },
    {"type": "TEXT", "name": "current_device_type", "value": "desktop", "scope": "LIFE_TIME"},
    {"type": "TEXT", "name": "current_platform", "value": "MacOs", "scope": "LIFE_TIME"},
    {"type": "TEXT", "scope": "LIFE_TIME", "name": "generated_by", "value": "USER"},
]

# ==================== TEST SHEET SCHEMA ====================
# In-memory rows are 0-indexed lists; spreadsheet row = in-memory index + 2 (header is row 1).
TS_COL_QUERY = 0
TS_COL_TEST_ID = 1
TS_COL_PAYLOAD = 2
TS_COL_CONV_ID = 3
TS_COL_USER_ID = 4
TS_COL_LOGIN_ID = 5
TS_COL_RESPONSE = 6
TS_NUM_COLS = 7
TS_HEADER_DEFAULT = ["Query", "TestID", "Payload", "conversationId", "userId", "LoginID", "Responses"]

# ==================== QUERIES SHEET SCHEMA (read-only from this runner's perspective) ====================
Q_COL_QUERY = 0
Q_COL_LOGIN_ID = 1
Q_COL_PASSWORD = 2
Q_COL_CONV_ID = 3
Q_COL_USER_ID = 4
Q_NUM_COLS = 5

SKIPPED_MARKER = "SKIPPED: prior turn failed"


# ==================== xlsx I/O ====================
def _stringify_cell(v: Any) -> str:
    """Normalize a cell value to a trimmed-friendly string."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, float):
        # openpyxl often returns numeric IDs (e.g. LoginIDs) as floats
        if v.is_integer():
            return str(int(v))
        return str(v)
    return str(v)


def load_xlsx(path: Path) -> Tuple[List[List[str]], List[str]]:
    """Read an xlsx. Returns (data_rows, header). Header is row 1."""
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    # Some xlsx files (especially ones generated outside Excel) carry a malformed
    # "active sheet" reference. openpyxl strips it and wb.active returns None.
    # Fall back to the first worksheet in that case.
    if ws is None:
        sheets = wb.worksheets
        if not sheets:
            wb.close()
            raise RuntimeError(f"{path} has no worksheets")
        ws = sheets[0]
    rows: List[List[str]] = []
    header: List[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        row_list = [_stringify_cell(c) for c in row]
        if i == 0:
            header = row_list
        else:
            rows.append(row_list)
    wb.close()
    return rows, header


def _save_xlsx_once(path: Path, header: List[str], rows: List[List[str]]) -> None:
    """Single atomic save attempt."""
    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for r in rows:
        ws.append(r)

    fd, tmp_path = tempfile.mkstemp(
        prefix=path.stem + "_", suffix=".tmp", dir=str(path.parent)
    )
    os.close(fd)  # openpyxl opens by path; close the mkstemp handle first
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, path)
    except Exception:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise
    finally:
        wb.close()


def save_xlsx(path: Path, header: List[str], rows: List[List[str]]) -> None:
    """
    Atomic xlsx save with retry on transient IO errors.

    If the target xlsx is briefly locked (e.g. user has it open in Excel
    and Excel/the OS holds an exclusive write fd for a moment), the
    os.replace can hit PermissionError/OSError. We retry with short
    backoff so a sub-second lock contention window doesn't surface as a
    hard failure. After the retry budget the original error is raised.

    On macOS/APFS, simply having the file open in Excel rarely causes
    contention (atomic rename succeeds anyway). On Windows, mandatory
    locking is the norm and the retries cover most "file in use" cases
    as long as the user closes the file within ~5 seconds.
    """
    max_retries = 5
    backoff_s = 0.5
    last_err: Optional[Exception] = None
    for attempt in range(max_retries + 1):
        try:
            _save_xlsx_once(path, header, rows)
            return
        except (PermissionError, OSError) as e:
            if isinstance(e, FileNotFoundError):
                raise  # not a lock contention issue
            last_err = e
            if attempt < max_retries:
                wait = backoff_s * (attempt + 1)
                print(
                    f"[WARN] save_xlsx({path.name}) hit IO error on attempt "
                    f"{attempt + 1}/{max_retries + 1}: {e}; retrying in "
                    f"{wait:.1f}s. (If you have the file open in Excel and "
                    f"this keeps happening, close it.)",
                    file=sys.stderr,
                )
                time.sleep(wait)
    if last_err is not None:
        raise last_err


def normalize_test_sheet_header(header: List[str]) -> List[str]:
    out = list(header) if header else []
    while len(out) < TS_NUM_COLS:
        out.append(TS_HEADER_DEFAULT[len(out)])
    return out[:TS_NUM_COLS]


def pad_row(row: List[str], n: int) -> List[str]:
    while len(row) < n:
        row.append("")
    return row


# ==================== queries.xlsx → user pool ====================
def load_user_pool(queries_path: Path) -> List[Tuple[str, str, str]]:
    """
    Returns ordered list of (conversationId, userId, LoginID) triples for rows
    where D and E are both populated and non-ERROR. Row order is preserved.
    """
    rows, _ = load_xlsx(queries_path)
    pool: List[Tuple[str, str, str]] = []
    for r in rows:
        r = pad_row(list(r), Q_NUM_COLS)
        conv_id = (r[Q_COL_CONV_ID] or "").strip()
        user_id = (r[Q_COL_USER_ID] or "").strip()
        login_id = (r[Q_COL_LOGIN_ID] or "").strip()
        if not conv_id or not user_id:
            continue
        # Match our error marker shape ("ERROR: ...") with the trailing colon so
        # a legitimately captured ID that incidentally starts with the letters
        # E-R-R-O-R isn't filtered out.
        if conv_id.startswith("ERROR:") or user_id.startswith("ERROR:"):
            continue
        if conv_id == "OK" or user_id == "OK":
            continue
        pool.append((conv_id, user_id, login_id))
    return pool


# ==================== Test sheet → TestID groups ====================
def group_by_test_id(rows: List[List[str]]) -> "OrderedDict[str, List[dict]]":
    """
    Walk rows in file order, group by TestID preserving first-appearance order.
    Each value is a list of per-turn dicts with: row_index (0-based within rows),
    query, and the existing conv_id/user_id/login_id if already filled.
    Empty rows (blank Query or blank TestID) are silently skipped.
    """
    groups: "OrderedDict[str, List[dict]]" = OrderedDict()
    for i, r in enumerate(rows):
        padded = pad_row(list(r), TS_NUM_COLS)
        query = (padded[TS_COL_QUERY] or "").strip()
        test_id = (padded[TS_COL_TEST_ID] or "").strip()
        if not query or not test_id:
            continue
        if test_id not in groups:
            groups[test_id] = []
        groups[test_id].append({
            "row_index": i,
            "query": query,
            "conv_id": (padded[TS_COL_CONV_ID] or "").strip(),
            "user_id": (padded[TS_COL_USER_ID] or "").strip(),
            "login_id": (padded[TS_COL_LOGIN_ID] or "").strip(),
        })
    return groups


# ==================== Assignment ====================
def assign_users_sequential(
    groups: "OrderedDict[str, List[dict]]",
    user_pool: List[Tuple[str, str, str]],
) -> Tuple[Dict[str, Tuple[str, str, str]], List[str]]:
    """TestID k -> user_pool[k]. Overflow TestIDs (no user available) end up in skipped."""
    assignments: Dict[str, Tuple[str, str, str]] = {}
    skipped: List[str] = []
    for i, test_id in enumerate(groups.keys()):
        if i < len(user_pool):
            assignments[test_id] = user_pool[i]
        else:
            skipped.append(test_id)
    return assignments, skipped


def assign_users_manual(
    groups: "OrderedDict[str, List[dict]]",
    user_pool: List[Tuple[str, str, str]],
) -> Dict[str, Tuple[str, str, str]]:
    """
    Validates that every TestID has D/E/F filled and consistent across all turns,
    that each triple exists in queries.xlsx, and that no conversationId is shared
    between TestIDs. Raises RuntimeError on any violation.
    """
    pool_set = {(p[0], p[1], p[2]) for p in user_pool}
    pool_by_conv = {p[0]: p for p in user_pool}
    assignments: Dict[str, Tuple[str, str, str]] = {}
    used_conv_ids: Dict[str, str] = {}  # conv_id -> test_id

    for test_id, turns in groups.items():
        # Every turn must have all three fields populated.
        for t in turns:
            if not t["conv_id"] or not t["user_id"] or not t["login_id"]:
                raise RuntimeError(
                    f"Manual assignment error: TestID {test_id} row {t['row_index'] + 2}: "
                    f"some of columns D/E/F are blank. In manual mode every turn of every "
                    f"TestID must have D/E/F filled."
                )
        # Values must be consistent across all turns of one TestID.
        conv_ids = {t["conv_id"] for t in turns}
        user_ids = {t["user_id"] for t in turns}
        login_ids = {t["login_id"] for t in turns}
        if len(conv_ids) != 1 or len(user_ids) != 1 or len(login_ids) != 1:
            raise RuntimeError(
                f"Manual assignment error: TestID {test_id} has inconsistent D/E/F values "
                f"across turns. All rows of one TestID must share the same user."
            )
        c = conv_ids.pop()
        u = user_ids.pop()
        l = login_ids.pop()
        # Triple must exist in queries.xlsx.
        if (c, u, l) not in pool_set:
            pool_hit = pool_by_conv.get(c)
            if pool_hit is None:
                raise RuntimeError(
                    f"Manual assignment error: TestID {test_id} references conversationId "
                    f"'{c}' which does not exist in queries.xlsx."
                )
            raise RuntimeError(
                f"Manual assignment error: TestID {test_id} assignment "
                f"(conv={c}, user={u}, login={l}) does not match the queries.xlsx entry "
                f"for that conversationId (conv={pool_hit[0]}, user={pool_hit[1]}, login={pool_hit[2]})."
            )
        # Strict 1:1 lock.
        if c in used_conv_ids:
            raise RuntimeError(
                f"Manual assignment error: TestIDs {used_conv_ids[c]} and {test_id} both "
                f"use conversationId '{c}'. Each user can be assigned to at most one TestID."
            )
        used_conv_ids[c] = test_id
        assignments[test_id] = (c, u, l)

    return assignments


# ==================== API calls ====================
def _build_post_body(conversation_id: str, user_id: str, query: str) -> dict:
    return {
        "conversationId": conversation_id,
        "messagePayload": {
            "text": query,
            "label": query,
            "messageId": str(uuid.uuid4()),
            "timestamp": int(time.time() * 1000),
        },
        "additionalAttributes": {
            # Deep-copy the static list so concurrent workers can't mutate it.
            "CUSTOM_ATTRIBUTES": [dict(a) for a in POST_BODY_STATIC_CUSTOM_ATTRIBUTES],
        },
        "userDetails": {
            "userId": user_id,
            "emailId": "",
        },
        "attachmentList": [],
    }


async def _post_message(
    client: httpx.AsyncClient,
    conversation_id: str,
    user_id: str,
    query: str,
    timeout_s: float,
    bot_ref_id: str,
) -> str:
    """One POST. Returns the requestID string. Raises RuntimeError on any failure."""
    body = _build_post_body(conversation_id, user_id, query)
    # Build per-request headers by merging the static set with the active
    # bot_ref_id. Each environment uses a different bot reference ID, so this
    # MUST be set per call rather than baked into POST_HEADERS.
    request_headers = {**POST_HEADERS, "x-bot-ref-id": bot_ref_id}

    if DEBUG_DUMP_REQUESTS:
        print(f"\nPOST {POST_URL}", flush=True)
        print(f"x-bot-ref-id: {bot_ref_id}", flush=True)
        print("Payload:", flush=True)
        print(json.dumps(body, indent=2), flush=True)

    try:
        resp = await client.post(POST_URL, headers=request_headers, json=body, timeout=timeout_s)
    except Exception as e:
        raise RuntimeError(f"POST network error: {e}")

    if DEBUG_DUMP_REQUESTS:
        print("Response:", flush=True)
        print(resp.text, flush=True)

    if resp.status_code >= 400:
        raise RuntimeError(f"POST HTTP {resp.status_code}: {resp.text[:200]}")

    try:
        data = resp.json()
    except Exception:
        raise RuntimeError(f"POST returned non-JSON body: {resp.text[:200]}")

    if not isinstance(data, dict):
        raise RuntimeError(f"POST returned non-object body: {str(data)[:200]}")

    status_code = data.get("statusCode")
    if status_code != "SUCCESS":
        raise RuntimeError(
            f"POST statusCode={status_code!r} statusMessage={data.get('statusMessage')!r}"
        )

    payload = data.get("payload") or []
    if not isinstance(payload, list) or len(payload) == 0:
        raise RuntimeError(f"POST returned empty or non-list payload: {str(payload)[:200]}")

    request_id = payload[0]
    if not isinstance(request_id, str) or not request_id.strip():
        raise RuntimeError(f"POST payload[0] is not a non-empty string: {payload!r}")

    return request_id.strip()


# Compiled once at import time. Used by _clean_html_text below.
_BR_TAG_RE = re.compile(r"<br\s*/?>", re.IGNORECASE)
_HTML_TAG_RE = re.compile(r"<[^>]+>")


def _clean_html_text(text: str) -> str:
    """
    Convert the HTML-flavored text the bot sends into plain text:
      - <br> / <br/> / <br /> become newlines.
      - All other tags (e.g. <p>, <a>, <strong>) are stripped.
      - HTML entities (&amp;, &nbsp;, &quot;, etc.) are decoded.
    """
    text = _BR_TAG_RE.sub("\n", text)
    text = _HTML_TAG_RE.sub("", text)
    text = html.unescape(text)
    return text.strip()


# Per-attachment fields that may carry user-visible bot reply text. Order
# determines how we concatenate them within a single attachment (title first
# gives the user context, description elaborates, text is the body, etc.).
# Tighten or extend if a new attachment type ships with a different field name.
_ATTACHMENT_TEXT_FIELDS_IN_ORDER = ("title", "description", "text", "subtitle", "body")

# Field names we look for on quickReply.options[] entries (and similar
# button-list constructs). Different bot deployments label the visible text
# differently — try these in order until one yields a non-empty string.
_OPTION_TEXT_FIELD_CANDIDATES = ("text", "label", "title", "value", "displayText", "name")


def _extract_text_from_option(opt: Any) -> Optional[str]:
    """Pull text from a single button/option entry (str or dict)."""
    if isinstance(opt, str):
        s = opt.strip()
        return _clean_html_text(s) if s else None
    if isinstance(opt, dict):
        for field in _OPTION_TEXT_FIELD_CANDIDATES:
            v = opt.get(field)
            if isinstance(v, str) and v.strip():
                cleaned = _clean_html_text(v)
                if cleaned:
                    return cleaned
    return None


def _extract_quick_reply_options(att: dict) -> List[str]:
    """
    Pull user-visible text from quickReply.options[] on an attachment, if any.
    Each option may be a bare string or a dict with one of several text-like
    fields (text/label/title/etc. — see _OPTION_TEXT_FIELD_CANDIDATES). Returns
    an empty list if there are no options or none have extractable text.
    """
    qr = att.get("quickReply")
    if not isinstance(qr, dict):
        return []
    options = qr.get("options")
    if not isinstance(options, list) or not options:
        return []
    out: List[str] = []
    for opt in options:
        text = _extract_text_from_option(opt)
        if text:
            out.append(text)
    return out


def _extract_button_labels(att: dict) -> List[str]:
    """
    Pull button text from the common interactive-button arrays the widget
    can ship: top-level `buttons`, `actions`, `options`. Used for Card and
    Button-style attachments (separate from quickReply, which is its own thing).
    """
    out: List[str] = []
    for key in ("buttons", "actions"):
        items = att.get(key)
        if not isinstance(items, list):
            continue
        for item in items:
            text = _extract_text_from_option(item)
            if text:
                out.append(text)
    return out


def _extract_form_field_labels(att: dict) -> List[str]:
    """
    Pull visible field labels from a Form attachment's schema[]. The widget
    renders these as input-field captions (e.g. "New Credit Limit") that a
    manual user definitely sees.
    """
    schema = att.get("schema")
    if not isinstance(schema, list):
        return []
    out: List[str] = []
    for field in schema:
        if not isinstance(field, dict):
            continue
        # Some fields have a `labels` list (multi-line) and/or a single `label`.
        labels = field.get("labels")
        if isinstance(labels, list):
            for lbl in labels:
                if isinstance(lbl, str) and lbl.strip():
                    cleaned = _clean_html_text(lbl)
                    if cleaned:
                        out.append(cleaned)
        single = field.get("label")
        if isinstance(single, str) and single.strip():
            cleaned = _clean_html_text(single)
            if cleaned:
                out.append(cleaned)
    return out


def _extract_form_submit_label(att: dict) -> Optional[str]:
    """Pull a Form attachment's submit-button label (e.g. 'Submit')."""
    label = att.get("submitButtonLabel")
    if isinstance(label, str) and label.strip():
        cleaned = _clean_html_text(label)
        return cleaned or None
    return None


def _try_extract_reply(data: Any) -> Optional[str]:
    """
    Return the concatenated bot reply text if `data` contains one or more
    attachments with user-visible content, otherwise None.

    What we accept:
      - `data` is a non-empty list whose first element is a dict with a
        non-empty `responses` array.
      - For each response, every attachment is inspected. Any attachment
        with at least one non-empty string in its title/description/text/
        subtitle/body fields contributes to the output.

    Why no `type` or `attachmentResponseType` filtering:
      - The bot uses different `type` values for different reply formats:
            ai.msg.domain.responses.core.Text  → plain text reply
            ai.msg.domain.formBuilder.Form     → form-input prompt
            ai.msg.domain.responses.core.Card  → card display
            (and others not yet observed)
        A type filter would silently drop the unfamiliar shapes. We extract
        from any attachment that exposes user-visible text in one of the
        known field names.
      - `attachmentResponseType` similarly varies per intent family
        (ANSWER_AI_RESPONSE, INTENT_RESPONSE, KNOWLEDGE_RESPONSE, etc.) and
        filtering would lose valid replies.

    HTML tags inside any extracted field are stripped and entities are
    decoded for cleaner spreadsheet output.
    """
    if not isinstance(data, list) or len(data) == 0:
        return None
    first = data[0]
    if not isinstance(first, dict):
        return None
    responses = first.get("responses")
    if not isinstance(responses, list) or not responses:
        return None

    extracted: List[str] = []
    for response_obj in responses:
        if not isinstance(response_obj, dict):
            continue
        attachments = response_obj.get("attachments")
        if not isinstance(attachments, list) or not attachments:
            continue
        for att in attachments:
            if not isinstance(att, dict):
                continue
            parts: List[str] = []
            # 1. Plain text fields (title / description / text / subtitle / body).
            for field_name in _ATTACHMENT_TEXT_FIELDS_IN_ORDER:
                raw = att.get(field_name)
                if not isinstance(raw, str) or not raw.strip():
                    continue
                cleaned = _clean_html_text(raw)
                if cleaned:
                    parts.append(cleaned)
            # 2. Quick-reply button labels (tappable suggestions like
            #    "Yes" / "No" or "I don't see my card / I found it").
            qr_options = _extract_quick_reply_options(att)
            if qr_options:
                parts.append(f"Quick replies: {' | '.join(qr_options)}")
            # 3. Card / Button attachment buttons (separate from quickReply).
            button_labels = _extract_button_labels(att)
            if button_labels:
                parts.append(f"Buttons: {' | '.join(button_labels)}")
            # 4. Form attachment field labels (e.g. "New Credit Limit"
            #    that captions an input field on a Form).
            form_fields = _extract_form_field_labels(att)
            if form_fields:
                parts.append(f"Form fields: {' | '.join(form_fields)}")
            # 5. Form submit button text (e.g. "Submit").
            submit_label = _extract_form_submit_label(att)
            if submit_label:
                parts.append(f"Submit button: {submit_label}")
            if parts:
                # Within one attachment, separate sections with single newlines
                # (so the body, then quick replies, then buttons, then form
                # fields read as a tight block).
                extracted.append("\n".join(parts))

    if not extracted:
        return None
    # Between attachments, separate with a blank line so multi-attachment
    # replies don't run together visually.
    return "\n\n".join(extracted)


async def _fetch_reply(
    client: httpx.AsyncClient,
    request_id: str,
    max_attempts: int,
    retry_delay_s: float,
    timeout_s: float,
    log_prefix: str,
) -> str:
    """
    Suffix-flip polling: odd attempts use the bare request_id, even attempts
    append "_res". Returns the extracted reply text on the first desired
    response; raises RuntimeError if all attempts are exhausted.
    """
    last_reason = "no attempts made"
    # Track the most recent HTTP 200 response body so we can dump it on
    # exhaustion. If our extractor returns None despite a 200, the body
    # captured here is exactly what the server gave us — so you can paste
    # it and we'll know whether the response shape is something the
    # extractor isn't recognizing.
    last_200_body: Optional[str] = None
    for attempt in range(1, max_attempts + 1):
        suffix = "" if attempt % 2 == 1 else "_res"
        probe_id = request_id + suffix
        url = GET_URL_TEMPLATE.format(request_id=probe_id)

        if DEBUG_DUMP_REQUESTS:
            print(f"\nGET {url}", flush=True)

        # Include a wall-clock timestamp on each attempt log so the user can
        # verify the per-TestID polling gap is actually retry_delay_s. With
        # multiple TestIDs running in parallel the combined log stream looks
        # faster than the per-TestID rate, which can be misleading.
        ts = time.strftime("%H:%M:%S", time.localtime())

        try:
            resp = await client.get(url, timeout=timeout_s)
        except Exception as e:
            last_reason = f"attempt {attempt} network error: {e}"
            print(
                f"  [{ts}] {log_prefix} GET attempt {attempt}/{max_attempts} "
                f"(suffix={suffix or 'none'}): network error: {e}",
                file=sys.stderr,
            )
        else:
            if DEBUG_DUMP_REQUESTS:
                print("Response:", flush=True)
                try:
                    print(resp.text, flush=True)
                except Exception:
                    print("<could not decode response body>", flush=True)

            if resp.status_code != 200:
                last_reason = f"attempt {attempt} HTTP {resp.status_code}"
                print(
                    f"  [{ts}] {log_prefix} GET attempt {attempt}/{max_attempts} "
                    f"(suffix={suffix or 'none'}): HTTP {resp.status_code}",
                    file=sys.stderr,
                )
            else:
                try:
                    data = resp.json()
                except Exception:
                    last_reason = f"attempt {attempt} non-JSON body"
                    print(
                        f"  [{ts}] {log_prefix} GET attempt {attempt}/{max_attempts} "
                        f"(suffix={suffix or 'none'}): non-JSON body",
                        file=sys.stderr,
                    )
                else:
                    reply = _try_extract_reply(data)
                    if reply is not None:
                        return reply
                    last_reason = f"attempt {attempt} HTTP 200 but no extractable text"
                    last_200_body = resp.text
                    print(
                        f"  [{ts}] {log_prefix} GET attempt {attempt}/{max_attempts} "
                        f"(suffix={suffix or 'none'}): HTTP 200 but no extractable "
                        f"text yet — body snippet:",
                        file=sys.stderr,
                    )
                    # Print the raw body so future shape mismatches are visible
                    # in real time, not only on exhaustion. Truncated to keep
                    # logs manageable when polling many TestIDs.
                    print(
                        f"    {resp.text[:1200]}",
                        file=sys.stderr,
                    )

        if attempt < max_attempts:
            await asyncio.sleep(retry_delay_s)

    # Polling exhausted. Dump copy-pasteable curls for both URL variants (bare
    # requestId and `_res` suffix) so the caller can reproduce manually in
    # Postman / a terminal and figure out what the server is actually
    # returning at this point.
    bare_url = GET_URL_TEMPLATE.format(request_id=request_id)
    res_url = GET_URL_TEMPLATE.format(request_id=request_id + "_res")
    print(
        f"[FAIL] {log_prefix}: All {max_attempts} GET attempts exhausted. "
        f"Reproduce manually with:",
        file=sys.stderr,
    )
    print(
        f"  curl -X GET '{bare_url}' "
        f"-H 'User-Agent: {BROWSER_USER_AGENT}'",
        file=sys.stderr,
    )
    print(
        f"  curl -X GET '{res_url}' "
        f"-H 'User-Agent: {BROWSER_USER_AGENT}'",
        file=sys.stderr,
    )
    # Show the last HTTP 200 body our extractor rejected (if any). If Postman
    # returned actual bot content for this requestId while our script gave up,
    # this snippet shows exactly what we were looking at — and lets us adjust
    # the extractor for the response shape that's actually in flight.
    if last_200_body is not None:
        print(
            f"[FAIL] {log_prefix}: last HTTP 200 body our extractor rejected "
            f"(first 1500 chars):",
            file=sys.stderr,
        )
        print(f"  {last_200_body[:1500]}", file=sys.stderr)
    else:
        print(
            f"[FAIL] {log_prefix}: no HTTP 200 ever received during polling "
            f"(everything was 202 / network errors / non-JSON). The server "
            f"never published a response in our window.",
            file=sys.stderr,
        )
    raise RuntimeError(f"GET exhausted {max_attempts} attempts (last: {last_reason})")


class _GetPollingExhausted(RuntimeError):
    """
    Raised when GET polling exhausts all attempts but the POST that triggered
    it succeeded. Carries the still-pending request_id so the caller can
    decide to extend GET polling on the SAME requestId rather than redoing
    the POST (which would create a new requestId while the original's
    response could still arrive late).
    """
    def __init__(self, message: str, request_id: str):
        super().__init__(message)
        self.request_id = request_id


async def _send_turn(
    client: httpx.AsyncClient,
    conversation_id: str,
    user_id: str,
    query: str,
    turn_timeout_s: float,
    get_max_attempts: int,
    get_retry_delay_s: float,
    log_prefix: str,
    bot_ref_id: str,
) -> Tuple[str, str]:
    """
    One full turn: POST → extract requestID → GET polling → extract reply text.
    Returns (request_id, reply_text). Raises _GetPollingExhausted if POST
    succeeded but GET polling timed out (caller should extend polling on the
    same requestId). Raises plain RuntimeError on any other failure (POST-
    level error, etc.) — caller should retry the whole POST+GET cycle.
    """
    request_id = await _post_message(
        client, conversation_id, user_id, query, turn_timeout_s, bot_ref_id,
    )
    try:
        reply = await _fetch_reply(
            client, request_id, get_max_attempts, get_retry_delay_s, turn_timeout_s, log_prefix,
        )
    except RuntimeError as e:
        # _fetch_reply raises RuntimeError with a "GET exhausted ..." prefix
        # when polling runs out. Re-wrap to carry the pending request_id.
        if str(e).startswith("GET exhausted"):
            raise _GetPollingExhausted(str(e), request_id) from e
        raise
    return request_id, reply


# ==================== Per-TestID worker ====================
async def _run_test_id(
    client: httpx.AsyncClient,
    test_id: str,
    turns: List[dict],
    user_triple: Tuple[str, str, str],
    rows: List[List[str]],
    header: List[str],
    test_sheet_path: Path,
    save_lock: asyncio.Lock,
    turn_retries: int,
    turn_timeout_s: float,
    inter_turn_delay_s: float,
    get_max_attempts: int,
    get_retry_delay_s: float,
    stats: dict,
    bot_ref_id: str,
) -> None:
    conv_id, user_id, login_id = user_triple
    skip_rest = False
    total_turns = len(turns)
    had_failure = False

    for turn_idx, turn in enumerate(turns):
        row_index = turn["row_index"]
        query = turn["query"]
        log_prefix = f"[{test_id} {turn_idx + 1}/{total_turns}]"

        if skip_rest:
            async with save_lock:
                row = pad_row(rows[row_index], TS_NUM_COLS)
                row[TS_COL_CONV_ID] = conv_id
                row[TS_COL_USER_ID] = user_id
                row[TS_COL_LOGIN_ID] = login_id
                row[TS_COL_PAYLOAD] = SKIPPED_MARKER
                row[TS_COL_RESPONSE] = SKIPPED_MARKER
                rows[row_index] = row
                save_xlsx(test_sheet_path, header, rows)
            stats["skipped_turns"] += 1
            print(f"[SKIP] {log_prefix}: skipped due to prior failure", file=sys.stderr)
            continue

        # Turn-level retry loop: total attempts = turn_retries + 1 (initial + retries).
        # Two distinct retry behaviors based on what failed:
        #   - POST failed (network, server error, malformed response): redo
        #     POST+GET on the next attempt.
        #   - POST succeeded but GET polling exhausted: do NOT redo POST.
        #     Re-poll the SAME requestId. The bot may simply be slow; a new
        #     POST would create a new requestId while the first one's
        #     response could still arrive after our window.
        last_err: Optional[Exception] = None
        request_id: Optional[str] = None
        reply_text: Optional[str] = None
        pending_request_id: Optional[str] = None
        total_attempts = turn_retries + 1
        for attempt in range(1, total_attempts + 1):
            try:
                if pending_request_id is not None:
                    print(
                        f"[INFO] {log_prefix} attempt {attempt}/{total_attempts}: "
                        f"continuing GET polling on prior requestId="
                        f"{pending_request_id[:8]}... (no new POST)",
                        file=sys.stderr,
                    )
                    reply_text = await _fetch_reply(
                        client, pending_request_id,
                        get_max_attempts, get_retry_delay_s, turn_timeout_s, log_prefix,
                    )
                    request_id = pending_request_id
                else:
                    request_id, reply_text = await _send_turn(
                        client, conv_id, user_id, query,
                        turn_timeout_s, get_max_attempts, get_retry_delay_s, log_prefix,
                        bot_ref_id,
                    )
                break
            except _GetPollingExhausted as e:
                last_err = e
                pending_request_id = e.request_id
                if attempt < total_attempts:
                    print(
                        f"[WARN] {log_prefix} attempt {attempt}/{total_attempts}: "
                        f"GET polling exhausted on requestId="
                        f"{pending_request_id[:8]}...; will extend polling on "
                        f"same id next round",
                        file=sys.stderr,
                    )
                    await asyncio.sleep(1.5 * attempt)
                else:
                    print(f"[FAIL] {log_prefix}: {e}", file=sys.stderr)
            except RuntimeError as e:
                # POST-side error or _fetch_reply error before exhaustion. Need
                # a fresh POST on next attempt — clear the pending tag.
                last_err = e
                pending_request_id = None
                if attempt < total_attempts:
                    print(
                        f"[WARN] {log_prefix} attempt {attempt}/{total_attempts}: {e}",
                        file=sys.stderr,
                    )
                    await asyncio.sleep(1.5 * attempt)
                else:
                    print(f"[FAIL] {log_prefix}: {e}", file=sys.stderr)
            except Exception as e:
                # Catch-all so an unexpected error doesn't get silently raised
                # past the gather wrapper.
                last_err = e
                pending_request_id = None
                if attempt < total_attempts:
                    print(
                        f"[WARN] {log_prefix} attempt {attempt}/{total_attempts}: {e}",
                        file=sys.stderr,
                    )
                    await asyncio.sleep(1.5 * attempt)
                else:
                    print(f"[FAIL] {log_prefix}: {e}", file=sys.stderr)

        async with save_lock:
            row = pad_row(rows[row_index], TS_NUM_COLS)
            row[TS_COL_CONV_ID] = conv_id
            row[TS_COL_USER_ID] = user_id
            row[TS_COL_LOGIN_ID] = login_id
            if reply_text is not None:
                row[TS_COL_PAYLOAD] = request_id or ""
                row[TS_COL_RESPONSE] = reply_text
            else:
                err_str = f"ERROR: {last_err}"[:300]
                # Preserve the requestId in column C if the POST succeeded but GET failed.
                row[TS_COL_PAYLOAD] = request_id if request_id else err_str
                row[TS_COL_RESPONSE] = err_str
            rows[row_index] = row
            save_xlsx(test_sheet_path, header, rows)

        if reply_text is not None:
            preview = reply_text.replace("\n", " ")[:80]
            req_short = (request_id or "")[:8]
            print(f"[OK]   {log_prefix} req={req_short}... -> \"{preview}\"")
            stats["completed_turns"] += 1
            if turn_idx < total_turns - 1 and inter_turn_delay_s > 0:
                await asyncio.sleep(inter_turn_delay_s)
        else:
            skip_rest = True
            had_failure = True
            stats["failed_turns"] += 1

    if had_failure:
        stats["partial_test_ids"] += 1
    else:
        stats["completed_test_ids"] += 1


# ==================== Main orchestrator ====================
async def _run_with_sem(
    sem: asyncio.Semaphore,
    client: httpx.AsyncClient,
    test_id: str,
    turns: List[dict],
    triple: Tuple[str, str, str],
    rows: List[List[str]],
    header: List[str],
    test_sheet_path: Path,
    save_lock: asyncio.Lock,
    turn_retries: int,
    turn_timeout_s: float,
    inter_turn_delay_s: float,
    get_max_attempts: int,
    get_retry_delay_s: float,
    stats: dict,
    bot_ref_id: str,
) -> None:
    async with sem:
        await _run_test_id(
            client, test_id, turns, triple, rows, header, test_sheet_path, save_lock,
            turn_retries, turn_timeout_s, inter_turn_delay_s,
            get_max_attempts, get_retry_delay_s, stats, bot_ref_id,
        )


async def run_api_flow(
    queries_path: Path,
    test_sheet_path: Path,
    workers: int,
    manual_assignment: bool,
    turn_retries: int,
    turn_timeout_s: float,
    inter_turn_delay_s: float,
    get_max_attempts: int,
    get_retry_delay_s: float,
    skip_done: bool = False,
    bot_ref_id: Optional[str] = None,
) -> int:
    if bot_ref_id is None:
        bot_ref_id = DEFAULT_BOT_REF_ID
    print("=" * 72)
    print("API FLOW: ATBFinancial multi-turn conversation runner")
    print(f"  queries:       {queries_path}")
    print(f"  test sheet:    {test_sheet_path}")
    print(f"  workers:       {workers}")
    print(f"  manual assign: {manual_assignment}")
    print(f"  turn retries:  {turn_retries}  (total attempts per turn: {turn_retries + 1})")
    print(f"  GET attempts:  {get_max_attempts}  (delay {get_retry_delay_s}s)")
    print(f"  turn timeout:  {turn_timeout_s}s")
    print(f"  bot_ref_id:    {bot_ref_id}")
    print("=" * 72)

    if not queries_path.exists():
        print(f"[ERROR] queries workbook not found: {queries_path}", file=sys.stderr)
        return 2
    if not test_sheet_path.exists():
        print(f"[ERROR] test sheet workbook not found: {test_sheet_path}", file=sys.stderr)
        return 2

    # Load user pool
    user_pool = load_user_pool(queries_path)
    print(f"Loaded {len(user_pool)} valid user(s) from {queries_path.name}")
    if not user_pool:
        print(
            "[ERROR] No valid users in queries workbook. Run the UI flow first to populate "
            "columns D (conversationId) and E (userId).",
            file=sys.stderr,
        )
        return 2

    # Load test sheet
    rows, header = load_xlsx(test_sheet_path)
    header = normalize_test_sheet_header(header)
    for i in range(len(rows)):
        rows[i] = pad_row(rows[i], TS_NUM_COLS)

    groups = group_by_test_id(rows)
    if not groups:
        print(
            "[ERROR] No valid TestID groups found in test sheet. Every row needs a non-empty "
            "Query (column A) and TestID (column B).",
            file=sys.stderr,
        )
        return 2

    total_turns_input = sum(len(v) for v in groups.values())
    print(f"Loaded {len(groups)} TestID group(s), {total_turns_input} total turn(s)")

    # Skip TestIDs that already have a non-error Response in column G. Used by
    # the pipelined orchestrator to avoid re-processing TestIDs that were
    # already handled in an earlier batch. Also collect the user_ids assigned
    # to those done TestIDs and exclude them from user_pool below — otherwise
    # sequential assignment would re-pair the same user with another TestID
    # (1:1 lock violation across batches).
    used_user_ids: set = set()
    if skip_done and groups:
        before = len(groups)
        for test_id in list(groups.keys()):
            turns = groups[test_id]
            # A TestID is considered "done" if all of its turns have a
            # non-empty Response cell that doesn't look like one of OUR
            # error markers. We match "ERROR:" and "SKIPPED:" with the
            # trailing colon so a legitimate bot response that just happens
            # to start with the word "Error" or "Skipped" isn't misclassified.
            done = True
            for t in turns:
                row = rows[t["row_index"]]
                resp = (row[TS_COL_RESPONSE] if len(row) > TS_COL_RESPONSE else "") or ""
                resp = resp.strip()
                if (not resp) or resp.startswith("ERROR:") or resp.startswith("SKIPPED:"):
                    done = False
                    break
            if done:
                # Capture the userId stored on this completed TestID so we
                # don't re-assign the same user to another TestID below.
                first_row = rows[turns[0]["row_index"]]
                u = (first_row[TS_COL_USER_ID] if len(first_row) > TS_COL_USER_ID else "") or ""
                u = u.strip()
                if u:
                    used_user_ids.add(u)
                del groups[test_id]
        skipped_done = before - len(groups)
        if skipped_done:
            print(f"skip_done=True — skipped {skipped_done} TestID(s) already completed")
        if not groups:
            print("All TestIDs already done; nothing to do.")
            return 0
        if used_user_ids:
            before_pool = len(user_pool)
            user_pool = [u for u in user_pool if u[1] not in used_user_ids]
            excluded = before_pool - len(user_pool)
            if excluded:
                print(
                    f"skip_done=True — excluded {excluded} user(s) already "
                    f"assigned to completed TestIDs (preserving 1:1 lock)"
                )

    # Assignment
    skipped_test_ids: List[str] = []
    if manual_assignment:
        try:
            assignments = assign_users_manual(groups, user_pool)
        except RuntimeError as e:
            print(f"[ERROR] {e}", file=sys.stderr)
            return 2
        print(f"Manual assignment validated: {len(assignments)} TestID(s) ready to run")
    else:
        assignments, skipped_test_ids = assign_users_sequential(groups, user_pool)
        print(
            f"Sequential assignment: {len(assignments)} TestID(s) paired, "
            f"{len(skipped_test_ids)} overflow"
        )
        if skipped_test_ids:
            print("[WARN] Not enough users for these TestIDs — they will be left untouched:")
            for t in skipped_test_ids:
                print(f"         - {t}")

    # Pre-write D/E/F in sequential mode so the file is self-documenting before turns start.
    if not manual_assignment and assignments:
        for test_id, triple in assignments.items():
            for turn in groups[test_id]:
                row = pad_row(rows[turn["row_index"]], TS_NUM_COLS)
                row[TS_COL_CONV_ID] = triple[0]
                row[TS_COL_USER_ID] = triple[1]
                row[TS_COL_LOGIN_ID] = triple[2]
                rows[turn["row_index"]] = row
        save_xlsx(test_sheet_path, header, rows)

    # Run TestIDs in parallel (semaphore-gated); turns within each TestID are sequential.
    stats = {
        "completed_test_ids": 0,
        "partial_test_ids": 0,
        "completed_turns": 0,
        "failed_turns": 0,
        "skipped_turns": 0,
    }
    sem = asyncio.Semaphore(workers)
    save_lock = asyncio.Lock()

    timeout_cfg = httpx.Timeout(turn_timeout_s, connect=min(30.0, turn_timeout_s))
    # Default the User-Agent at the client level so GET requests (which don't
    # pass through POST_HEADERS) also carry the browser-shaped UA.
    async with httpx.AsyncClient(
        timeout=timeout_cfg,
        headers={"User-Agent": BROWSER_USER_AGENT},
    ) as client:
        tasks = [
            asyncio.create_task(_run_with_sem(
                sem, client, test_id, groups[test_id], triple, rows, header,
                test_sheet_path, save_lock,
                turn_retries, turn_timeout_s, inter_turn_delay_s,
                get_max_attempts, get_retry_delay_s, stats, bot_ref_id,
            ))
            for test_id, triple in assignments.items()
        ]
        results = await asyncio.gather(*tasks, return_exceptions=True)
        for test_id, res in zip(assignments.keys(), results):
            if isinstance(res, Exception):
                print(
                    f"[FAIL] TestID {test_id} worker raised unhandled exception: {res}",
                    file=sys.stderr,
                )

    # Summary
    print("=" * 72)
    print("API FLOW SUMMARY")
    print(f"  TestIDs fully completed:        {stats['completed_test_ids']}")
    print(f"  TestIDs partial (turn failed):  {stats['partial_test_ids']}")
    print(f"  TestIDs skipped (no user):      {len(skipped_test_ids)}")
    print(f"  Turns completed:                {stats['completed_turns']}")
    print(f"  Turns failed:                   {stats['failed_turns']}")
    print(f"  Turns skipped after failure:    {stats['skipped_turns']}")
    print("=" * 72)

    if stats["failed_turns"] == 0 and not skipped_test_ids:
        return 0
    return 1


async def run_retry_failed(
    test_sheet_path: Path,
    workers: int,
    turn_retries: int,
    turn_timeout_s: float,
    inter_turn_delay_s: float,
    get_max_attempts: int,
    get_retry_delay_s: float,
    bot_ref_id: Optional[str] = None,
) -> int:
    """
    Re-run TestIDs that previously failed.

    A TestID is "previously failed" if any of its turns has an ERROR:- or
    SKIPPED:-prefixed Response (column G). For each such TestID, we read the
    user that was originally assigned (columns D/E/F), clear the failure
    markers from G (and from C if Payload also held an ERROR), and re-run the
    full POST+GET pipeline using that same user. TestIDs whose D/E columns
    are blank are skipped — there's no original user to retry with.

    Doesn't touch TestIDs that are fully completed or never started.
    """
    if bot_ref_id is None:
        bot_ref_id = DEFAULT_BOT_REF_ID
    print("=" * 72)
    print(f"API RETRY: scanning {test_sheet_path.name} for failed TestIDs")
    print(f"  bot_ref_id: {bot_ref_id}")
    print("=" * 72)

    if not test_sheet_path.exists():
        print(f"[ERROR] Test sheet not found: {test_sheet_path}", file=sys.stderr)
        return 2

    rows, header = load_xlsx(test_sheet_path)
    header = normalize_test_sheet_header(header)
    for i in range(len(rows)):
        rows[i] = pad_row(rows[i], TS_NUM_COLS)

    groups = group_by_test_id(rows)
    if not groups:
        print("No TestID groups in test sheet.")
        return 0

    failed_groups: "OrderedDict[str, List[dict]]" = OrderedDict()
    assignments: Dict[str, Tuple[str, str, str]] = {}
    skipped_no_user: List[str] = []

    # Track original turn count per TestID so we can log a meaningful
    # "retrying turns N–M of TOTAL" line.
    original_turn_count: Dict[str, int] = {}

    for test_id, turns in groups.items():
        # Find the FIRST turn whose Response is one of OUR error markers.
        # Important: we re-run only from this turn onward, leaving earlier
        # successful turns untouched. Re-running them would (a) regress those
        # responses if the retry hit a transient failure, and (b) pollute the
        # bot's conversation memory by re-sending messages it already
        # answered.
        first_failed_idx = -1
        for i, t in enumerate(turns):
            row = rows[t["row_index"]]
            resp = (row[TS_COL_RESPONSE] if len(row) > TS_COL_RESPONSE else "") or ""
            resp = resp.strip()
            if resp.startswith("ERROR:") or resp.startswith("SKIPPED:"):
                first_failed_idx = i
                break
        if first_failed_idx == -1:
            continue  # No failures in this TestID; leave it alone.
        # Pull the original assignment from the first turn's row (D/E/F is
        # pre-written for every turn by run_api_flow, so any turn would do).
        first_row = rows[turns[0]["row_index"]]
        conv_id = (first_row[TS_COL_CONV_ID] if len(first_row) > TS_COL_CONV_ID else "").strip()
        user_id = (first_row[TS_COL_USER_ID] if len(first_row) > TS_COL_USER_ID else "").strip()
        login_id = (first_row[TS_COL_LOGIN_ID] if len(first_row) > TS_COL_LOGIN_ID else "").strip()
        if not conv_id or not user_id:
            skipped_no_user.append(test_id)
            continue
        original_turn_count[test_id] = len(turns)
        failed_groups[test_id] = turns[first_failed_idx:]
        assignments[test_id] = (conv_id, user_id, login_id)

    if not failed_groups:
        print("No failed TestIDs to retry.")
        if skipped_no_user:
            print(
                f"[WARN] {len(skipped_no_user)} TestID(s) had failures but no "
                f"user in D/E — skipped: {skipped_no_user}"
            )
        return 0

    print(f"Found {len(failed_groups)} failed TestID(s) to retry:")
    for tid in failed_groups:
        triple = assignments[tid]
        retry_count = len(failed_groups[tid])
        total = original_turn_count[tid]
        first_idx = total - retry_count + 1  # 1-based for human readability
        print(
            f"  - {tid}  user={triple[1]}  retrying turns {first_idx}-{total} "
            f"of {total}  login={triple[2] or '(none)'}"
        )
    if skipped_no_user:
        print(
            f"[WARN] Skipping {len(skipped_no_user)} TestID(s) with failures "
            f"but no original user in D/E: {skipped_no_user}"
        )

    # Wipe ERROR:/SKIPPED: markers so the retry can write fresh values.
    # Keep D/E/F intact (those are the original user we're reusing).
    for test_id, turns in failed_groups.items():
        for turn in turns:
            row = pad_row(rows[turn["row_index"]], TS_NUM_COLS)
            resp = (row[TS_COL_RESPONSE] or "").strip()
            if resp.startswith("ERROR:") or resp.startswith("SKIPPED:"):
                row[TS_COL_RESPONSE] = ""
            payload = (row[TS_COL_PAYLOAD] or "").strip()
            if payload.startswith("ERROR:") or payload.startswith("SKIPPED:"):
                row[TS_COL_PAYLOAD] = ""
            rows[turn["row_index"]] = row
    save_xlsx(test_sheet_path, header, rows)

    # Now run the same per-TestID orchestration as run_api_flow, with these
    # specific assignments.
    stats = {
        "completed_test_ids": 0,
        "partial_test_ids": 0,
        "completed_turns": 0,
        "failed_turns": 0,
        "skipped_turns": 0,
    }
    sem = asyncio.Semaphore(workers)
    save_lock = asyncio.Lock()

    timeout_cfg = httpx.Timeout(turn_timeout_s, connect=min(30.0, turn_timeout_s))
    async with httpx.AsyncClient(
        timeout=timeout_cfg,
        headers={"User-Agent": BROWSER_USER_AGENT},
    ) as client:
        tasks = [
            asyncio.create_task(_run_with_sem(
                sem, client, test_id, failed_groups[test_id], triple, rows, header,
                test_sheet_path, save_lock,
                turn_retries, turn_timeout_s, inter_turn_delay_s,
                get_max_attempts, get_retry_delay_s, stats, bot_ref_id,
            ))
            for test_id, triple in assignments.items()
        ]
        results = await asyncio.gather(*tasks, return_exceptions=True)
        for test_id, res in zip(assignments.keys(), results):
            if isinstance(res, Exception):
                print(
                    f"[FAIL] Retry of TestID {test_id} raised unhandled "
                    f"exception: {res}",
                    file=sys.stderr,
                )

    print("=" * 72)
    print("API RETRY SUMMARY")
    print(f"  TestIDs retried:                {len(failed_groups)}")
    print(f"  TestIDs fully completed:        {stats['completed_test_ids']}")
    print(f"  TestIDs partial (turn failed):  {stats['partial_test_ids']}")
    print(f"  TestIDs without user (skipped): {len(skipped_no_user)}")
    print(f"  Turns completed:                {stats['completed_turns']}")
    print(f"  Turns failed:                   {stats['failed_turns']}")
    print(f"  Turns skipped after failure:    {stats['skipped_turns']}")
    print("=" * 72)

    return 0 if stats["failed_turns"] == 0 else 1


def retry_failed_sync(
    test_sheet_path: str = DEFAULT_TEST_SHEET_XLSX,
    workers: int = DEFAULT_WORKERS,
    turn_retries: int = DEFAULT_TURN_RETRIES,
    turn_timeout_s: float = DEFAULT_TURN_TIMEOUT_S,
    inter_turn_delay_s: float = DEFAULT_INTER_TURN_DELAY_S,
    get_max_attempts: int = DEFAULT_GET_MAX_ATTEMPTS,
    get_retry_delay_s: float = DEFAULT_GET_RETRY_DELAY_S,
    bot_ref_id: Optional[str] = None,
) -> int:
    """
    Synchronous entry point for the retry-failed pass. Callable from the UI
    script's main() after the regular pipeline finishes.
    """
    return asyncio.run(run_retry_failed(
        Path(test_sheet_path).resolve(),
        workers,
        turn_retries,
        turn_timeout_s,
        inter_turn_delay_s,
        get_max_attempts,
        get_retry_delay_s,
        bot_ref_id,
    ))


def run_sync(
    queries_path: str = DEFAULT_QUERIES_XLSX,
    test_sheet_path: str = DEFAULT_TEST_SHEET_XLSX,
    workers: int = DEFAULT_WORKERS,
    manual_assignment: bool = DEFAULT_MANUAL_ASSIGNMENT,
    turn_retries: int = DEFAULT_TURN_RETRIES,
    turn_timeout_s: float = DEFAULT_TURN_TIMEOUT_S,
    inter_turn_delay_s: float = DEFAULT_INTER_TURN_DELAY_S,
    get_max_attempts: int = DEFAULT_GET_MAX_ATTEMPTS,
    get_retry_delay_s: float = DEFAULT_GET_RETRY_DELAY_S,
    skip_done: bool = False,
    bot_ref_id: Optional[str] = None,
) -> int:
    """
    Synchronous entry point for use by netomi_conversation_extractor.py after its
    UI flow completes, or from the standalone __main__ block below.
    """
    return asyncio.run(run_api_flow(
        Path(queries_path).resolve(),
        Path(test_sheet_path).resolve(),
        workers,
        manual_assignment,
        turn_retries,
        turn_timeout_s,
        inter_turn_delay_s,
        get_max_attempts,
        get_retry_delay_s,
        skip_done,
        bot_ref_id,
    ))


if __name__ == "__main__":
    # Standalone usage: re-run just the API flow against existing xlsx files.
    # All tuning via the DEFAULT_* constants above. Do not invoke this concurrently
    # with netomi_conversation_extractor.py — they share the xlsx files.
    try:
        rc = run_sync()
    except KeyboardInterrupt:
        print("Interrupted.", file=sys.stderr)
        rc = 130
    sys.exit(rc)